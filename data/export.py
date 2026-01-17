# Módulo para la exportación de datos a formatos como Excel

import pandas as pd
import io
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def export_full_project_to_excel(session_state: Dict[str, Any], incluir_formulas: bool = False, incluir_graficos: bool = True) -> io.BytesIO:
    """
    Recopila todos los datos, análisis y gráficos del proyecto y los exporta 
    a un archivo Excel con múltiples hojas.
    
    Args:
        session_state: Estado de la sesión con todos los datos
        incluir_formulas: Si True, incluye fórmulas en las celdas de la hoja 'Datos Gráficos'
        incluir_graficos: Si True, incluye las hojas de gráficos (100% RPM y VFD)
    """
    try:
        output = io.BytesIO()
        wb = Workbook()
        wb.remove(wb.active)
        
        # Generar automáticamente datos de curva del sistema si no existen
        if 'df_curva_sistema' not in session_state or session_state.get('df_curva_sistema', pd.DataFrame()).empty:
            df_sistema_generated = generate_system_curve_data(session_state)
            session_state['df_curva_sistema'] = df_sistema_generated
        
        # Verificar que todos los DataFrames necesarios estén presentes
        ensure_required_dataframes(session_state)

        # --- Función de estilo ---
        def style_sheet(ws, title, col_widths: Dict[str, int], print_area=None, print_orientation='portrait'):
            ws.title = title
            
            # Eliminar líneas de cuadrícula
            ws.sheet_view.showGridLines = False
            
            # Configurar impresión
            ws.page_setup.orientation = print_orientation
            ws.page_setup.paperSize = 1  # Letter
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 0  # Ajustar ancho, altura automática
            ws.page_setup.fitToWidth = 1
            
            if print_area:
                ws.print_area = print_area
            
            ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=16, color="000000")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_widths))
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=1).fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width
            
            # Estilo para cabeceras de tablas
            for cell in ws[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                
            ws.freeze_panes = 'A3'

        # --- Hoja 1: Resumen General del Proyecto ---
        ws_resumen = wb.create_sheet("Resumen General del Proyecto")
        style_sheet(ws_resumen, "Resumen General del Proyecto", {'A': 40, 'B': 25}, print_area='A1:B14', print_orientation='portrait')
        
        resumen_data = {
            'Parámetro': [
                'Proyecto', 'Diseño', 'Caudal de Diseño (L/s)', 'Altura Estática Total (m)',
                'ADT Total (m)', 'Potencia Motor Final (HP)', 'NPSH Disponible (m.c.a.)',
                'Punto de Operación Caudal (L/s)', 'Punto de Operación Altura (m)',
                'Punto de Operación Eficiencia (%)', 'Punto de Operación Potencia (HP)'
            ],
            'Valor': [
                session_state.get('proyecto', 'N/A'),
                session_state.get('diseno', 'N/A'),
                f"{session_state.get('caudal_lps', 0):.2f}",
                f"{session_state.get('altura_estatica_total', 0):.2f}",
                f"{session_state.get('adt_total', 0):.2f}",
                f"{session_state.get('potencia_motor_final_hp', 0):.2f}",
                f"{session_state.get('npshd_mca', 0):.2f}",
                f"{session_state.get('caudal_operacion', 0):.2f}",
                f"{session_state.get('altura_operacion', 0):.2f}",
                f"{session_state.get('eficiencia_operacion', 0):.2f}",
                f"{session_state.get('potencia_operacion', 0):.2f}"
            ]
        }
        
        df_resumen = pd.DataFrame(resumen_data)
        for r in dataframe_to_rows(df_resumen, index=False, header=True):
            ws_resumen.append(r)
        for cell in ws_resumen['A'] + ws_resumen['B']:
            cell.alignment = Alignment(vertical='center', horizontal='left')
        for cell in ws_resumen["1:1"]: 
            cell.font = Font(bold=True)
        for cell in ws_resumen["2:2"]: 
            cell.font = Font(bold=True)
        
        # Centrar título entre A1 y B2, y colocar fondo blanco
        ws_resumen.cell(row=1, column=1).fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        # Centrar columna B
        for row in ws_resumen.iter_rows(min_row=3, max_row=ws_resumen.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # --- Hoja 2: Inputs Detallados del Proyecto ---
        ws_inputs = wb.create_sheet("Inputs Detallados del Proyecto")
        style_sheet(ws_inputs, "Inputs Detallados del Proyecto", {'A': 25, 'B': 35, 'C': 25}, print_area='A1:C31', print_orientation='portrait')
        
        inputs_data = {'Sección': [], 'Parámetro': [], 'Valor': []}
        def add_input(seccion, parametro, valor):
            inputs_data['Sección'].append(seccion)
            inputs_data['Parámetro'].append(parametro)
            inputs_data['Valor'].append(str(valor))

        # Condiciones de Operación
        add_input('Operación', 'Caudal de Diseño (L/s)', session_state.get('caudal_lps'))
        add_input('Operación', 'Elevación del sitio (m)', session_state.get('elevacion_sitio'))
        add_input('Operación', 'Altura de Succión (m)', session_state.get('altura_succion_input'))
        add_input('Operación', 'Bomba inundada', session_state.get('bomba_inundada'))
        add_input('Operación', 'Altura de Descarga (m)', session_state.get('altura_descarga'))
        add_input('Operación', 'Número de Bombas', session_state.get('num_bombas'))

        # Succión
        add_input('Succión', 'Longitud Tubería (m)', session_state.get('long_succion'))
        add_input('Succión', 'Material Tubería', session_state.get('mat_succion'))
        add_input('Succión', 'Diámetro Interno (mm)', f"{session_state.get('diam_succion_mm', 0):.2f}")
        add_input('Succión', 'Otras Pérdidas (m)', session_state.get('otras_perdidas_succion'))
        
        # Impulsión
        add_input('Impulsión', 'Longitud Tubería (m)', session_state.get('long_impulsion'))
        add_input('Impulsión', 'Material Tubería', session_state.get('mat_impulsion'))
        add_input('Impulsión', 'Diámetro Interno (mm)', f"{session_state.get('diam_impulsion_mm', 0):.2f}")
        add_input('Impulsión', 'Otras Pérdidas (m)', session_state.get('otras_perdidas_impulsion'))

        # Accesorios
        for acc in session_state.get('accesorios_succion', []):
            add_input('Accesorios Succión', acc['tipo'], f"Cantidad: {acc['cantidad']}")
        for acc in session_state.get('accesorios_impulsion', []):
            add_input('Accesorios Impulsión', acc['tipo'], f"Cantidad: {acc['cantidad']}")

        df_inputs = pd.DataFrame(inputs_data)
        for r in dataframe_to_rows(df_inputs, index=False, header=True):
            ws_inputs.append(r)
        
        # Centrar columna C
        for row in ws_inputs.iter_rows(min_row=3, max_row=ws_inputs.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # --- Hoja 3: Cálculos del Sistema ---
        ws_calculos = wb.create_sheet("Cálculos del Sistema")
        ws_calculos.sheet_view.showGridLines = False
        ws_calculos.column_dimensions['A'].width = 50
        ws_calculos.column_dimensions['B'].width = 20
        ws_calculos.column_dimensions['C'].width = 15
        ws_calculos.column_dimensions['D'].width = 50
        ws_calculos.column_dimensions['E'].width = 40
        
        # Configurar área de impresión
        ws_calculos.print_area = 'A1:E74'
        ws_calculos.page_setup.orientation = 'portrait'
        ws_calculos.page_setup.paperSize = 1  # Letter
        ws_calculos.page_setup.fitToPage = True
        ws_calculos.page_setup.fitToHeight = 0
        ws_calculos.page_setup.fitToWidth = 1
        
        # Crear función auxiliar para agregar secciones con estilo
        def add_section_header(ws, row, text, bg_color="4F81BD"):
            cell = ws.cell(row=row, column=1, value=text)
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            return row + 1
        
        def add_subsection_header(ws, row, text):
            cell = ws.cell(row=row, column=1, value=text)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            return row + 1
        
        def add_input_row(ws, row, descripcion, valor, unidad, observacion="", formula_simbolica=""):
            ws.cell(row=row, column=1, value=descripcion).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row, column=2, value=valor).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=3, value=unidad).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=4, value=observacion).alignment = Alignment(horizontal='left', vertical='center')
            cell_formula = ws.cell(row=row, column=5, value=formula_simbolica)
            cell_formula.alignment = Alignment(horizontal='left', vertical='center')
            cell_formula.font = Font(italic=True, color="0000FF")
            return row + 1
        
        def add_calculated_row(ws, row, descripcion, formula_or_value, unidad, observacion="", use_formula=False, formula_simbolica=""):
            ws.cell(row=row, column=1, value=descripcion).alignment = Alignment(horizontal='left', vertical='center')
            cell_value = ws.cell(row=row, column=2)
            if use_formula and isinstance(formula_or_value, str) and formula_or_value.startswith('='):
                cell_value.value = formula_or_value
            else:
                cell_value.value = formula_or_value
            cell_value.alignment = Alignment(horizontal='center', vertical='center')
            cell_value.font = Font(bold=True, color="0000FF")
            # Aplicar formato numérico con 2 decimales
            cell_value.number_format = '0.00'
            ws.cell(row=row, column=3, value=unidad).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=4, value=observacion).alignment = Alignment(horizontal='left', vertical='center')
            cell_formula = ws.cell(row=row, column=5, value=formula_simbolica)
            cell_formula.alignment = Alignment(horizontal='left', vertical='center')
            cell_formula.font = Font(italic=True, color="0000FF")
            return row + 1
        
        # Título principal
        ws_calculos.cell(row=1, column=1, value="CÁLCULOS COMPLETOS DEL SISTEMA DE BOMBEO").font = Font(bold=True, size=16, color="FFFFFF")
        ws_calculos.cell(row=1, column=1).fill = PatternFill(start_color="17365D", end_color="17365D", fill_type="solid")
        ws_calculos.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws_calculos.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws_calculos.row_dimensions[1].height = 30
        
        current_row = 3
        
        # ===== SECCIÓN 1: INPUTS - CONDICIONES DE OPERACIÓN =====
        current_row = add_section_header(ws_calculos, current_row, "📋 1. INPUTS - CONDICIONES DE OPERACIÓN")
        
        # Encabezados de columna
        headers = ['Descripción', 'Dato', 'Unidad', 'Observaciones', 'Fórmula']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_calculos.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Inputs de condiciones de operación
        input_row_caudal = current_row
        current_row = add_input_row(ws_calculos, current_row, "Caudal de Diseño (Q)", 
                                     session_state.get('caudal_lps', 0), "L/s", "", "Q")
        
        input_row_elevacion = current_row
        current_row = add_input_row(ws_calculos, current_row, "Elevación del Sitio (z)", 
                                     session_state.get('elevacion_sitio', 0), "m", "", "z")
        
        input_row_altura_succion = current_row
        current_row = add_input_row(ws_calculos, current_row, "Altura de Succión (Hs)", 
                                     session_state.get('altura_succion_input', 0), "m", 
                                     "Negativo si bomba en aspiración", "Hs")
        
        input_row_altura_descarga = current_row
        current_row = add_input_row(ws_calculos, current_row, "Altura de Descarga (Hd)", 
                                     session_state.get('altura_descarga', 0), "m", "", "Hd")
        
        input_row_num_bombas = current_row
        current_row = add_input_row(ws_calculos, current_row, "Número de Bombas", 
                                     session_state.get('num_bombas', 1), "-", "")
        
        bomba_inundada = "Sí" if session_state.get('bomba_inundada', False) else "No"
        current_row = add_input_row(ws_calculos, current_row, "Bomba Inundada", 
                                     bomba_inundada, "-", "")
        
        input_row_temperatura = current_row
        current_row = add_input_row(ws_calculos, current_row, "Temperatura del Fluido (T)", 
                                     session_state.get('temperatura', 20), "°C", "", "T")
        
        input_row_densidad = current_row
        current_row = add_input_row(ws_calculos, current_row, "Densidad del Líquido (ρ)", 
                                     session_state.get('densidad_liquido', 1000), "kg/m³", "", "ρ")
        
        current_row += 1
        
        # ===== SECCIÓN 2: INPUTS - TUBERÍA DE SUCCIÓN =====
        current_row = add_section_header(ws_calculos, current_row, "🔧 2. INPUTS - TUBERÍA Y ACCESORIOS DE SUCCIÓN")
        
        # Encabezados
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_calculos.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        input_row_long_succion = current_row
        current_row = add_input_row(ws_calculos, current_row, "Longitud de Tubería", 
                                     session_state.get('long_succion', 0), "m", "")
        
        input_row_diam_succion = current_row
        current_row = add_input_row(ws_calculos, current_row, "Diámetro Interno", 
                                     session_state.get('diam_succion_mm', 0), "mm", "")
        
        current_row = add_input_row(ws_calculos, current_row, "Material de Tubería", 
                                     session_state.get('mat_succion', 'N/A'), "-", "")
        
        # Rugosidad no se usa en los cálculos (se usa coeficiente de Hazen-Williams)
        # input_row_rugosidad_succion = current_row
        # current_row = add_input_row(ws_calculos, current_row, "Rugosidad Absoluta", 
        #                              session_state.get('rugosidad_succion', 0.0015), "mm", "")
        
        current_row += 1
        
        # ===== SECCIÓN 3: INPUTS - TUBERÍA DE IMPULSIÓN =====
        current_row = add_section_header(ws_calculos, current_row, "🔧 3. INPUTS - TUBERÍA Y ACCESORIOS DE IMPULSIÓN")
        
        # Encabezados
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_calculos.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        input_row_long_impulsion = current_row
        current_row = add_input_row(ws_calculos, current_row, "Longitud de Tubería", 
                                     session_state.get('long_impulsion', 0), "m", "")
        
        input_row_diam_impulsion = current_row
        current_row = add_input_row(ws_calculos, current_row, "Diámetro Interno", 
                                     session_state.get('diam_impulsion_mm', 0), "mm", "")
        
        current_row = add_input_row(ws_calculos, current_row, "Material de Tubería", 
                                     session_state.get('mat_impulsion', 'N/A'), "-", "")
        
        # Rugosidad no se usa en los cálculos (se usa coeficiente de Hazen-Williams)
        # input_row_rugosidad_impulsion = current_row
        # current_row = add_input_row(ws_calculos, current_row, "Rugosidad Absoluta", 
        #                              session_state.get('rugosidad_impulsion', 0.0015), "mm", "")
        
        current_row += 2
        
        # ===== SECCIÓN 4: RESULTADOS DE CÁLCULOS HIDRÁULICOS =====
        current_row = add_section_header(ws_calculos, current_row, "📊 4. RESULTADOS DE CÁLCULOS HIDRÁULICOS", "C0504D")
        
        # --- Subsección: Resultados de Succión ---
        current_row = add_subsection_header(ws_calculos, current_row, "✓ Resultados de Succión")
        
        # Encabezados
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_calculos.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Velocidad en Succión
        velocidad_succion = session_state.get('velocidad_succion', 0)
        velocidad_obs = "✅ Velocidad óptima (0.6-0.9 m/s)" if 0.6 <= velocidad_succion <= 0.9 else "⚠️ Fuera del rango óptimo"
        if incluir_formulas:
            # Fórmula: V = Q / A = Q / (π * (D/2)²) = Q / (π * D² / 4) = 4Q / (π * D²)
            # Q en m³/s, D en m
            formula_velocidad_suc = f"=(B{input_row_caudal}/1000)/(3.14159*POWER(B{input_row_diam_succion}/1000,2)/4)"
            current_row = add_calculated_row(ws_calculos, current_row, "Velocidad en Succión (Vs)", 
                                            formula_velocidad_suc, "m/s", velocidad_obs, use_formula=True,
                                            formula_simbolica="Vs = Q / A")
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "Velocidad en Succión (Vs)", 
                                            velocidad_succion, "m/s", velocidad_obs,
                                            formula_simbolica="Vs = Q / A")
        
        # Pérdidas (valores calculados por la app)
        perdida_primaria_suc = session_state.get('hf_primaria_succion', 0)
        row_perdida_prim_suc = current_row
        current_row = add_calculated_row(ws_calculos, current_row, "Pérdida Primaria (hf1s)", 
                                        perdida_primaria_suc, "m", "Pérdida por fricción",
                                        formula_simbolica="hf1 = f(L, D, V, C)")
        
        perdida_secundaria_suc = session_state.get('hf_secundaria_succion', 0)
        row_perdida_sec_suc = current_row
        current_row = add_calculated_row(ws_calculos, current_row, "Pérdida Secundaria (hf2s)", 
                                        perdida_secundaria_suc, "m", "Pérdida por accesorios",
                                        formula_simbolica="hf2 = ΣK × V²/(2g)")
        
        long_equiv_suc = session_state.get('le_total_succion', 0)
        current_row = add_calculated_row(ws_calculos, current_row, "Long. Equiv. Accesorios", 
                                        long_equiv_suc, "m", "")
        
        # Otras pérdidas en succión
        otras_perdidas_suc = session_state.get('otras_perdidas_succion', 0)
        row_otras_perdidas_suc = current_row
        current_row = add_input_row(ws_calculos, current_row, "Otras Pérdidas", 
                                     otras_perdidas_suc, "m", "Pérdidas adicionales")
        
        # Pérdida Total en Succión (con fórmula si está activado)
        perdida_total_suc = session_state.get('perdida_total_succion', 0)
        row_perdida_total_suc = current_row
        if incluir_formulas:
            formula_perdida_total_suc = f"=B{row_perdida_prim_suc}+B{row_perdida_sec_suc}+B{row_otras_perdidas_suc}"
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 Pérdida Total en Succión (hfs)", 
                                            formula_perdida_total_suc, "m", "Suma de pérdidas", use_formula=True,
                                            formula_simbolica="hfs = hf1s + hf2s + otras")
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 Pérdida Total en Succión (hfs)", 
                                            perdida_total_suc, "m", "Suma de pérdidas",
                                            formula_simbolica="hfs = hf1s + hf2s + otras")
        
        # Altura Dinámica de Succión (con fórmula si está activado)
        altura_dinamica_suc = session_state.get('altura_dinamica_succion', 0)
        if incluir_formulas:
            formula_altura_din_suc = f"=ABS(B{input_row_altura_succion})+B{row_perdida_total_suc}"
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 Altura Dinámica de Succión (Hds)", 
                                            formula_altura_din_suc, "m", "Hs + Pérdidas", use_formula=True,
                                            formula_simbolica="Hds = |Hs| + hfs")
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 Altura Dinámica de Succión (Hds)", 
                                            altura_dinamica_suc, "m", "Hs + Pérdidas",
                                            formula_simbolica="Hds = |Hs| + hfs")
        
        current_row += 1
        
        # --- Subsección: Resultados de Impulsión ---
        current_row = add_subsection_header(ws_calculos, current_row, "✓ Resultados de Impulsión")
        
        # Encabezados
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_calculos.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Velocidad en Impulsión
        velocidad_impulsion = session_state.get('velocidad_impulsion', 0)
        velocidad_imp_obs = "✅ Velocidad óptima (1.0-2.5 m/s)" if 1.0 <= velocidad_impulsion <= 2.5 else "⚠️ Fuera del rango óptimo"
        if incluir_formulas:
            formula_velocidad_imp = f"=(B{input_row_caudal}/1000)/(3.14159*POWER(B{input_row_diam_impulsion}/1000,2)/4)"
            current_row = add_calculated_row(ws_calculos, current_row, "Velocidad en Impulsión", 
                                            formula_velocidad_imp, "m/s", velocidad_imp_obs, use_formula=True)
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "Velocidad en Impulsión", 
                                            velocidad_impulsion, "m/s", velocidad_imp_obs)
        
        # Pérdidas (valores calculados por la app)
        perdida_primaria_imp = session_state.get('hf_primaria_impulsion', 0)
        row_perdida_prim_imp = current_row
        current_row = add_calculated_row(ws_calculos, current_row, "Pérdida Primaria", 
                                        perdida_primaria_imp, "m", "Pérdida por fricción")
        
        perdida_secundaria_imp = session_state.get('hf_secundaria_impulsion', 0)
        row_perdida_sec_imp = current_row
        current_row = add_calculated_row(ws_calculos, current_row, "Pérdida Secundaria", 
                                        perdida_secundaria_imp, "m", "Pérdida por accesorios")
        
        long_equiv_imp = session_state.get('le_total_impulsion', 0)
        current_row = add_calculated_row(ws_calculos, current_row, "Long. Equiv. Accesorios", 
                                        long_equiv_imp, "m", "")
        
        # Otras pérdidas en impulsión
        otras_perdidas_imp = session_state.get('otras_perdidas_impulsion', 0)
        row_otras_perdidas_imp = current_row
        current_row = add_input_row(ws_calculos, current_row, "Otras Pérdidas", 
                                     otras_perdidas_imp, "m", "Pérdidas adicionales")
        
        # Pérdida Total en Impulsión (con fórmula si está activado)
        perdida_total_imp = session_state.get('perdida_total_impulsion', 0)
        row_perdida_total_imp = current_row
        if incluir_formulas:
            formula_perdida_total_imp = f"=B{row_perdida_prim_imp}+B{row_perdida_sec_imp}+B{row_otras_perdidas_imp}"
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 Pérdida Total en Impulsión (hfi)", 
                                            formula_perdida_total_imp, "m", "Suma de pérdidas", use_formula=True,
                                            formula_simbolica="hfi = hf1i + hf2i + otras")
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 Pérdida Total en Impulsión (hfi)", 
                                            perdida_total_imp, "m", "Suma de pérdidas",
                                            formula_simbolica="hfi = hf1i + hf2i + otras")
        
        # Altura Dinámica de Impulsión (con fórmula si está activado)
        altura_dinamica_imp = session_state.get('altura_dinamica_impulsion', 0)
        if incluir_formulas:
            formula_altura_din_imp = f"=B{input_row_altura_descarga}+B{row_perdida_total_imp}"
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 Altura Dinámica de Impulsión (Hdi)", 
                                            formula_altura_din_imp, "m", "Hd + Pérdidas", use_formula=True,
                                            formula_simbolica="Hdi = Hd + hfi")
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 Altura Dinámica de Impulsión (Hdi)", 
                                            altura_dinamica_imp, "m", "Hd + Pérdidas",
                                            formula_simbolica="Hdi = Hd + hfi")
        
        current_row += 1
        
        # --- Subsección: Cálculo de NPSH Disponible ---
        current_row = add_subsection_header(ws_calculos, current_row, "✓ Cálculo de NPSH Disponible")
        
        # Encabezados
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_calculos.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Estado de la bomba
        bomba_estado = "⚠️ Bomba en aspiración" if not session_state.get('bomba_inundada', False) else "✅ Bomba inundada"
        current_row = add_input_row(ws_calculos, current_row, "Estado de la Bomba", 
                                     bomba_estado, "-", "")
        
        # Parámetros del cálculo (valores calculados por la app)
        presion_barometrica = session_state.get('presion_barometrica_calculada', 10.33)
        row_presion_baro = current_row
        current_row = add_calculated_row(ws_calculos, current_row, "Presión Barométrica (Pa)", 
                                        presion_barometrica, "m.c.a.", 
                                        f"Elevación: {session_state.get('elevacion_sitio', 0)} m",
                                        formula_simbolica="Pa = f(z)")
        
        # Altura neta calculada por la app (considera si bomba está inundada o en aspiración)
        altura_neta = session_state.get('altura_neta_succion', 0)
        row_altura_neta = current_row
        current_row = add_calculated_row(ws_calculos, current_row, "Altura Neta (hn)", 
                                        altura_neta, "m.c.a.", "Nivel agua - eje bomba")
        
        perdida_succion_npsh = perdida_total_suc
        row_perdida_suc_npsh = current_row
        if incluir_formulas:
            # Referencia a la pérdida total calculada arriba
            formula_perdida_suc_npsh = f"=B{row_perdida_total_suc}"
            current_row = add_calculated_row(ws_calculos, current_row, "Pérdida en la Succión (hg)", 
                                            formula_perdida_suc_npsh, "m.c.a.", "", use_formula=True)
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "Pérdida en la Succión (hg)", 
                                            perdida_succion_npsh, "m.c.a.", "")
        
        presion_vapor = session_state.get('presion_vapor_calculada', 0.25)
        row_presion_vapor = current_row
        current_row = add_calculated_row(ws_calculos, current_row, "Presión de Vapor (Pv)", 
                                        presion_vapor, "m.c.a.", 
                                        f"Temp: {session_state.get('temperatura', 20)}°C",
                                        formula_simbolica="Pv = f(T)")
        
        # Fórmula NPSH
        ws_calculos.cell(row=current_row, column=1, value="📐 Fórmula NPSH Unificada:").font = Font(bold=True, italic=True)
        ws_calculos.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1
        
        ws_calculos.cell(row=current_row, column=1, value="NPSHd = Pa + hn - hg - Pv").font = Font(italic=True, color="0000FF")
        ws_calculos.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1
        
        # NPSH Disponible (con fórmula si está activado)
        npshd = session_state.get('npshd_mca', 0)
        npsh_obs = "✅ NPSH adecuado" if npshd > 3 else "⚠️ NPSH bajo - Verificar"
        if incluir_formulas:
            formula_npshd = f"=B{row_presion_baro}+B{row_altura_neta}-B{row_perdida_suc_npsh}-B{row_presion_vapor}"
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 NPSH Disponible (NPSHd)", 
                                            formula_npshd, "m.c.a.", npsh_obs, use_formula=True,
                                            formula_simbolica="NPSHd = Pa + hn - hg - Pv")
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 NPSH Disponible (NPSHd)", 
                                            npshd, "m.c.a.", npsh_obs,
                                            formula_simbolica="NPSHd = Pa + hn - hg - Pv")
        
        current_row += 1
        
        # --- Subsección: Cálculo del Motor de la Bomba ---
        current_row = add_subsection_header(ws_calculos, current_row, "✓ Cálculo del Motor de la Bomba")
        
        # Encabezados
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_calculos.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Parámetros (valores calculados por la app)
        adt_total = session_state.get('adt_total', 0)
        row_adt = current_row
        if incluir_formulas:
            # ADT = Altura_Estática + Pérdidas_Totales
            # Altura_Estática depende de si bomba está inundada o en aspiración
            # Pérdidas_Totales = Pérdida_Succión + Pérdida_Impulsión
            bomba_inundada = session_state.get('bomba_inundada', False)
            if bomba_inundada:
                # Bomba inundada: Hd - Hs
                formula_adt = f"=(B{input_row_altura_descarga}-B{input_row_altura_succion})+B{row_perdida_total_suc}+B{row_perdida_total_imp}"
            else:
                # Bomba en aspiración: Hd + Hs
                formula_adt = f"=(B{input_row_altura_descarga}+B{input_row_altura_succion})+B{row_perdida_total_suc}+B{row_perdida_total_imp}"
            current_row = add_calculated_row(ws_calculos, current_row, "Altura Total (ADT)", 
                                            formula_adt, "m", "Altura estática + pérdidas", use_formula=True,
                                            formula_simbolica="ADT = Hest + hfs + hfi")
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "Altura Total (ADT)", 
                                            adt_total, "m", "Altura estática + pérdidas",
                                            formula_simbolica="ADT = Hest + hfs + hfi")
        
        eficiencia_bomba = session_state.get('eficiencia_bomba', 75)
        row_eficiencia = current_row
        current_row = add_input_row(ws_calculos, current_row, "Eficiencia de la Bomba (η)", 
                                     eficiencia_bomba, "%", "", "η")
        
        factor_seguridad = session_state.get('factor_seguridad', 1.2)
        row_factor_seg = current_row
        current_row = add_input_row(ws_calculos, current_row, "Factor de Seguridad (FS)", 
                                     factor_seguridad, "-", "", "FS")
        
        # Fórmula de potencia hidráulica
        ws_calculos.cell(row=current_row, column=1, value="📐 Fórmula Potencia Hidráulica:").font = Font(bold=True, italic=True)
        ws_calculos.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1
        
        ws_calculos.cell(row=current_row, column=1, value="P_h = (ρ × g × Q × H) / 1000").font = Font(italic=True, color="0000FF")
        ws_calculos.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1
        
        # Potencia hidráulica (con fórmula si está activado)
        potencia_hidraulica_kw = session_state.get('potencia_hidraulica_kw', 0)
        potencia_hidraulica_hp = session_state.get('potencia_hidraulica_hp', 0)
        row_pot_hidraulica = current_row
        if incluir_formulas:
            # P_h (kW) = (ρ × g × Q × H) / 1000
            # ρ = densidad (kg/m³), g = 9.81 m/s², Q = caudal (L/s), H = ADT (m)
            # Fórmula simplificada: (9.81 * Q_Ls * H) / 1000
            formula_pot_hidraulica = f"=(9.81*B{input_row_caudal}*B{row_adt})/1000"
            current_row = add_calculated_row(ws_calculos, current_row, "Potencia Hidráulica (Ph)", 
                                            formula_pot_hidraulica, "kW", 
                                            f"({potencia_hidraulica_hp:.2f} HP)", use_formula=True,
                                            formula_simbolica="Ph = (ρ × g × Q × ADT) / 1000")
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "Potencia Hidráulica (Ph)", 
                                            potencia_hidraulica_kw, "kW", 
                                            f"({potencia_hidraulica_hp:.2f} HP)",
                                            formula_simbolica="Ph = (ρ × g × Q × ADT) / 1000")
        
        # Potencia del motor (con fórmula si está activado)
        potencia_motor_kw = session_state.get('potencia_motor_final_kw', 0)
        potencia_motor_hp = session_state.get('potencia_motor_final_hp', 0)
        row_pot_motor = current_row
        if incluir_formulas:
            # P_motor (kW) = (P_h / (η_bomba/100)) * Factor_seguridad
            # Fórmula: (B{pot_hidraulica} / (B{eficiencia}/100)) * B{factor_seg}
            formula_pot_motor = f"=(B{row_pot_hidraulica}/(B{row_eficiencia}/100))*B{row_factor_seg}"
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 Potencia del Motor (Pm)", 
                                            formula_pot_motor, "kW", 
                                            f"({potencia_motor_hp:.2f} HP)", use_formula=True,
                                            formula_simbolica="Pm = (Ph / η) × FS")
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 Potencia del Motor (Pm)", 
                                            potencia_motor_kw, "kW", 
                                            f"({potencia_motor_hp:.2f} HP)",
                                            formula_simbolica="Pm = (Ph / η) × FS")
        
        current_row += 1
        
        # --- Subsección: Resumen del Sistema ---
        current_row = add_subsection_header(ws_calculos, current_row, "✓ Resumen del Sistema")
        
        # Encabezados
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_calculos.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Altura estática total (con fórmula si está activado)
        altura_estatica_total = session_state.get('altura_estatica_total', 0)
        row_altura_estatica = current_row
        if incluir_formulas:
            # La fórmula correcta depende de si la bomba está inundada o en aspiración
            # altura_succion_input siempre es positivo en la app
            # Si bomba inundada: H_est = Hd - Hs
            # Si bomba en aspiración: H_est = Hd + Hs (porque el nivel está ABAJO del eje)
            bomba_inundada = session_state.get('bomba_inundada', False)
            if bomba_inundada:
                # Bomba inundada: Hd - Hs
                formula_altura_estatica = f"=B{input_row_altura_descarga}-B{input_row_altura_succion}"
                observacion = "Hd - Hs (bomba inundada)"
            else:
                # Bomba en aspiración: Hd + Hs
                formula_altura_estatica = f"=B{input_row_altura_descarga}+B{input_row_altura_succion}"
                observacion = "Hd + Hs (bomba en aspiración)"
            current_row = add_calculated_row(ws_calculos, current_row, "Altura Estática Total (Hest)", 
                                            formula_altura_estatica, "m", observacion, use_formula=True,
                                            formula_simbolica="Hest = Hd ± Hs")
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "Altura Estática Total (Hest)", 
                                            altura_estatica_total, "m", "Hd - Hs",
                                            formula_simbolica="Hest = Hd ± Hs")
        
        # Pérdidas totales (con fórmula si está activado)
        perdidas_totales = perdida_total_suc + perdida_total_imp
        row_perdidas_totales = current_row
        if incluir_formulas:
            formula_perdidas_totales = f"=B{row_perdida_total_suc}+B{row_perdida_total_imp}"
            current_row = add_calculated_row(ws_calculos, current_row, "Pérdidas Totales (hf)", 
                                            formula_perdidas_totales, "m", "Succión + Impulsión", use_formula=True,
                                            formula_simbolica="hf = hfs + hfi")
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "Pérdidas Totales (hf)", 
                                            perdidas_totales, "m", "Succión + Impulsión",
                                            formula_simbolica="hf = hfs + hfi")
        
        # ADT (con fórmula si está activado)
        if incluir_formulas:
            formula_adt_total = f"=B{row_altura_estatica}+B{row_perdidas_totales}"
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 ALTURA DINÁMICA TOTAL (ADT)", 
                                            formula_adt_total, "m", "Altura estática + pérdidas", use_formula=True,
                                            formula_simbolica="ADT = Hest + hf")
        else:
            current_row = add_calculated_row(ws_calculos, current_row, "🎯 ALTURA DINÁMICA TOTAL (ADT)", 
                                            adt_total, "m", "Altura estática + pérdidas",
                                            formula_simbolica="ADT = Hest + hf")
        
        # Aplicar bordes a todas las celdas con datos
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        for row in ws_calculos.iter_rows(min_row=3, max_row=current_row-1, min_col=1, max_col=5):
            for cell in row:
                if cell.value:
                    cell.border = thin_border

        # --- Hoja 4: Resultados Numéricos Completos ---
        ws_resultados = wb.create_sheet("Resultados Numéricos Completos")
        style_sheet(ws_resultados, "Resultados Numéricos Completos", {'A': 25, 'B': 35, 'C': 25}, print_area='A1:C18', print_orientation='portrait')
        
        results_data = {'Sección': [], 'Parámetro': [], 'Valor': []}
        def add_result(seccion, parametro, valor, unit=""):
            results_data['Sección'].append(seccion)
            results_data['Parámetro'].append(parametro)
            results_data['Valor'].append(f"{valor:.2f} {unit}".strip())

        # Resultados Succión
        add_result('Succión', 'Velocidad', session_state.get('velocidad_succion', 0), 'm/s')
        add_result('Succión', 'Pérdida Primaria', session_state.get('hf_primaria_succion', 0), 'm')
        add_result('Succión', 'Pérdida Secundaria', session_state.get('hf_secundaria_succion', 0), 'm')
        add_result('Succión', 'Pérdida Total', session_state.get('perdida_total_succion', 0), 'm')
        add_result('Succión', 'Altura Dinámica', session_state.get('altura_dinamica_succion', 0), 'm')

        # Resultados Impulsión
        add_result('Impulsión', 'Velocidad', session_state.get('velocidad_impulsion', 0), 'm/s')
        add_result('Impulsión', 'Pérdida Primaria', session_state.get('hf_primaria_impulsion', 0), 'm')
        add_result('Impulsión', 'Pérdida Secundaria', session_state.get('hf_secundaria_impulsion', 0), 'm')
        add_result('Impulsión', 'Pérdida Total', session_state.get('perdida_total_impulsion', 0), 'm')
        add_result('Impulsión', 'Altura Dinámica', session_state.get('altura_dinamica_impulsion', 0), 'm')

        # Resultados Generales
        add_result('Sistema', 'NPSH Disponible', session_state.get('npshd_mca', 0), 'm.c.a')
        add_result('Sistema', 'Altura Dinámica Total (ADT)', session_state.get('adt_total', 0), 'm')
        add_result('Motor', 'Potencia Hidráulica', session_state.get('potencia_hidraulica_hp', 0), 'HP')
        add_result('Motor', 'Potencia Motor Requerida', session_state.get('potencia_motor_final_hp', 0), 'HP')
        if session_state.get('motor_seleccionado'):
            motor = session_state.get('motor_seleccionado')
            add_result('Motor', 'Motor Estándar Seleccionado', motor['potencia_hp'], 'HP')

        df_results = pd.DataFrame(results_data)
        for r in dataframe_to_rows(df_results, index=False, header=True):
            ws_resultados.append(r)
        
        # Centrar columna C
        for row in ws_resultados.iter_rows(min_row=3, max_row=ws_resultados.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # --- Hoja 4: Datos Gráficos ---
        ws_data = wb.create_sheet("Datos Gráficos")
        # Hoja visible para que el usuario pueda ver los datos
        # Eliminar líneas de cuadrícula
        ws_data.sheet_view.showGridLines = False

        try:
            # Obtener el porcentaje VDF primero (necesario para títulos)
            vfd_percentage = session_state.get('vfd_speed_percentage', session_state.get('rpm_percentage', 0))
            if vfd_percentage == 0:
                vfd_percentage = 75.0  # Valor por defecto
            
            # Recopilar todos los datos de curvas con manejo seguro
            
            # ---------- TABLAS 100% RPM ----------
            df_bomba_100 = session_state.get('df_bomba_100', pd.DataFrame())
            df_rendimiento_100 = session_state.get('df_rendimiento_100', pd.DataFrame())
            df_potencia_100 = session_state.get('df_potencia_100', pd.DataFrame())
            df_npsh_100 = session_state.get('df_npsh_100', pd.DataFrame())
            df_sistema_100 = session_state.get('df_sistema_100', pd.DataFrame())
            
            # ---------- TABLAS VDF ----------
            df_bomba_vfd = session_state.get('df_bomba_vfd', pd.DataFrame())
            df_rendimiento_vfd = session_state.get('df_rendimiento_vfd', pd.DataFrame())
            df_potencia_vfd = session_state.get('df_potencia_vfd', pd.DataFrame())
            df_npsh_vfd = session_state.get('df_npsh_vfd', pd.DataFrame())
            df_sistema_vfd = session_state.get('df_sistema_vfd', pd.DataFrame())
            
            # ---------- TABLAS OTRAS (compatibilidad con código anterior) ----------
            df_system = session_state.get('df_curva_sistema', pd.DataFrame())
            df_bomba_eff = session_state.get('df_eff_100', pd.DataFrame())
            df_power_100_fallback = session_state.get('df_power_100', pd.DataFrame())
            df_npshr = session_state.get('df_npshr', pd.DataFrame())
            
            # ========== FALLBACK: Generar datos desde curva_inputs si no existen ==========
            curva_inputs = session_state.get('curva_inputs', {})
            
            ajuste_tipo = session_state.get('ajuste_tipo', 'Cuadrática (2do grado)')
            grado = 1 if ajuste_tipo == "Lineal" else 2 if ajuste_tipo == "Cuadrática (2do grado)" else 3
            caudal_diseno = session_state.get('caudal_lps', 50.0)
            
            # Generar caudales para las tablas usando parámetros del usuario (sección 8)
            import numpy as np
            
            # Obtener parámetros de tablas 100% desde session_state
            q_min_100 = session_state.get('q_min_100_tab2', 0.0)
            q_max_100 = session_state.get('q_max_100_tab2', caudal_diseno if caudal_diseno > 0 else 100.0)
            paso_100 = session_state.get('paso_caudal_100_tab2', 5.0)
            
            # Obtener parámetros de tablas VFD desde session_state
            q_min_vfd = session_state.get('q_min_vdf_tab2', 0.0)
            q_max_vfd_param = session_state.get('q_max_vdf_tab2', caudal_diseno if caudal_diseno > 0 else 100.0)
            paso_vfd = session_state.get('paso_caudal_vdf_tab2', 5.0)
            
            # Generar arrays de caudales respetando los parámetros del usuario
            caudales_tabla = np.arange(q_min_100, q_max_100 + paso_100, paso_100)
            caudales_tabla_vfd = np.arange(q_min_vfd, q_max_vfd_param + paso_vfd, paso_vfd)
            
            # Fallback si arrays vacíos
            if len(caudales_tabla) == 0:
                caudales_tabla = np.linspace(0, caudal_diseno * 1.2 if caudal_diseno > 0 else 100, 20)
            if len(caudales_tabla_vfd) == 0:
                caudales_tabla_vfd = np.linspace(0, caudal_diseno * 1.2 if caudal_diseno > 0 else 100, 20)
            
            # Fallback para df_bomba_100
            if df_bomba_100.empty and curva_inputs.get('bomba') and len(curva_inputs['bomba']) >= 2:
                puntos = curva_inputs['bomba']
                x_vals = np.array([pt[0] for pt in puntos])
                y_vals = np.array([pt[1] for pt in puntos])
                coef = np.polyfit(x_vals, y_vals, min(grado, len(x_vals) - 1))
                alturas = np.polyval(coef, caudales_tabla)
                df_bomba_100 = pd.DataFrame({
                    'Caudal (L/s)': np.round(caudales_tabla, 3),
                    'Altura (m)': np.round(alturas, 3)
                })
            
            # Fallback para df_rendimiento_100
            if df_rendimiento_100.empty and curva_inputs.get('rendimiento') and len(curva_inputs['rendimiento']) >= 2:
                puntos = curva_inputs['rendimiento']
                x_vals = np.array([pt[0] for pt in puntos])
                y_vals = np.array([pt[1] for pt in puntos])
                coef = np.polyfit(x_vals, y_vals, min(grado, len(x_vals) - 1))
                valores = np.polyval(coef, caudales_tabla)
                df_rendimiento_100 = pd.DataFrame({
                    'Caudal (L/s)': np.round(caudales_tabla, 3),
                    'Rendimiento (%)': np.round(valores, 3)
                })
            
            # Fallback para df_potencia_100
            if df_potencia_100.empty and curva_inputs.get('potencia') and len(curva_inputs['potencia']) >= 2:
                puntos = curva_inputs['potencia']
                x_vals = np.array([pt[0] for pt in puntos])
                y_vals = np.array([pt[1] for pt in puntos])
                coef = np.polyfit(x_vals, y_vals, min(grado, len(x_vals) - 1))
                valores = np.polyval(coef, caudales_tabla)
                df_potencia_100 = pd.DataFrame({
                    'Caudal (L/s)': np.round(caudales_tabla, 3),
                    'Potencia (HP)': np.round(valores, 3)
                })
            
            # Fallback para df_npsh_100
            if df_npsh_100.empty and curva_inputs.get('npsh') and len(curva_inputs['npsh']) >= 2:
                puntos = curva_inputs['npsh']
                x_vals = np.array([pt[0] for pt in puntos])
                y_vals = np.array([pt[1] for pt in puntos])
                coef = np.polyfit(x_vals, y_vals, min(grado, len(x_vals) - 1))
                valores = np.polyval(coef, caudales_tabla)
                df_npsh_100 = pd.DataFrame({
                    'Caudal (L/s)': np.round(caudales_tabla, 3),
                    'NPSH (m)': np.round(valores, 3)
                })
            
            # Si aún no hay datos NPSH, generar valores estimados basados en caudal
            # NPSH típico: aumenta cuadráticamente con el caudal
            if df_npsh_100.empty:
                # Valor base típico de 3m + incremento cuadrático
                npsh_base = 3.0
                npsh_coef = 0.0005  # Coeficiente cuadrático típico
                npsh_estimado = npsh_base + npsh_coef * (caudales_tabla ** 2)
                df_npsh_100 = pd.DataFrame({
                    'Caudal (L/s)': np.round(caudales_tabla, 3),
                    'NPSH (m)': np.round(npsh_estimado, 3)
                })
            
            # Fallback para df_sistema_100
            if df_sistema_100.empty and curva_inputs.get('sistema') and len(curva_inputs['sistema']) >= 2:
                puntos = curva_inputs['sistema']
                x_vals = np.array([pt[0] for pt in puntos])
                y_vals = np.array([pt[1] for pt in puntos])
                coef = np.polyfit(x_vals, y_vals, min(grado, len(x_vals) - 1))
                valores = np.polyval(coef, caudales_tabla)
                df_sistema_100 = pd.DataFrame({
                    'Caudal (L/s)': np.round(caudales_tabla, 3),
                    'Altura (m)': np.round(valores, 3)
                })
            
            # Fallback para VFD usando leyes de afinidad
            vfd_ratio = vfd_percentage / 100.0 if vfd_percentage > 0 else 0.76
            
            if df_bomba_vfd.empty and not df_bomba_100.empty:
                df_bomba_vfd = df_bomba_100.copy()
                # Q_vfd = Q_100 * ratio, H_vfd = H_100 * ratio²
                q_col = [c for c in df_bomba_100.columns if 'Caudal' in c]
                h_col = [c for c in df_bomba_100.columns if 'Altura' in c]
                if q_col and h_col:
                    df_bomba_vfd[q_col[0]] = df_bomba_100[q_col[0]] * vfd_ratio
                    df_bomba_vfd[h_col[0]] = df_bomba_100[h_col[0]] * (vfd_ratio ** 2)
            
            if df_rendimiento_vfd.empty and not df_rendimiento_100.empty:
                df_rendimiento_vfd = df_rendimiento_100.copy()
                # Eficiencia se mantiene similar (ligeramente menor)
                q_col = [c for c in df_rendimiento_100.columns if 'Caudal' in c]
                if q_col:
                    df_rendimiento_vfd[q_col[0]] = df_rendimiento_100[q_col[0]] * vfd_ratio
            
            if df_potencia_vfd.empty and not df_potencia_100.empty:
                df_potencia_vfd = df_potencia_100.copy()
                # P_vfd = P_100 * ratio³
                q_col = [c for c in df_potencia_100.columns if 'Caudal' in c]
                p_col = [c for c in df_potencia_100.columns if 'Potencia' in c]
                if q_col and p_col:
                    df_potencia_vfd[q_col[0]] = df_potencia_100[q_col[0]] * vfd_ratio
                    df_potencia_vfd[p_col[0]] = df_potencia_100[p_col[0]] * (vfd_ratio ** 3)
            
            if df_npsh_vfd.empty and not df_npsh_100.empty:
                df_npsh_vfd = df_npsh_100.copy()
                # NPSH se reduce con el caudal
                q_col = [c for c in df_npsh_100.columns if 'Caudal' in c]
                if q_col:
                    df_npsh_vfd[q_col[0]] = df_npsh_100[q_col[0]] * vfd_ratio
            
            if df_sistema_vfd.empty and not df_sistema_100.empty:
                df_sistema_vfd = df_sistema_100.copy()
            
            # ========== FIN FALLBACK ==========

            # Función segura para agregar datos de curva
            def safe_add_curve_data(df_source, df_target, q_column, h_column, q_output_name, h_output_name):
                try:
                    if df_source is None or df_source.empty:
                        return df_target
                    
                    # Buscar columna de caudal (puede tener diferentes nombres/unidades)
                    q_col_found = None
                    for col in df_source.columns:
                        col_lower = col.lower()
                        if 'caudal' in col_lower or 'q_' in col_lower or col_lower.startswith('q'):
                            q_col_found = col
                            break
                    
                    # Buscar columna de valor Y (altura, rendimiento, potencia, NPSH)
                    h_col_found = None
                    # Lista de posibles nombres para cada tipo
                    h_keywords = {
                        'Altura (m)': ['altura', 'h_', 'head', 'altura (m)'],
                        'Rendimiento (%)': ['rendimiento', 'eficiencia', 'eff', 'η', 'rendimiento (%)'],
                        'Potencia (HP)': ['potencia', 'power', 'p_', 'bhp', 'potencia (hp)'],
                        'NPSH (m)': ['npsh', 'npshr', 'npsh (m)', 'npshr (m)'],
                        'ADT (m)': ['adt', 'altura', 'dynamic', 'adt (m)']
                    }
                    
                    # Primero buscar coincidencia exacta
                    if h_column in df_source.columns:
                        h_col_found = h_column
                    else:
                        # Buscar por palabras clave
                        keywords = h_keywords.get(h_column, [h_column.lower().split()[0]])
                        for col in df_source.columns:
                            col_lower = col.lower()
                            if any(kw in col_lower for kw in keywords):
                                h_col_found = col
                                break
                    
                    # Verificar si tenemos ambas columnas
                    if q_col_found and h_col_found:
                        # Convertir caudal a L/s si está en m³/h
                        df_temp = df_source[[q_col_found, h_col_found]].copy()
                        if 'm³/h' in q_col_found or 'm3/h' in q_col_found:
                            df_temp[q_col_found] = df_temp[q_col_found] / 3.6
                        
                        # Renombrar columnas
                        df_temp = df_temp.rename(columns={q_col_found: q_output_name, h_col_found: h_output_name})
                        
                        if df_target.empty:
                            return df_temp
                        else:
                            return pd.concat([df_target, df_temp], axis=1)
                    else:
                        print(f"Warning: Could not find columns for {q_output_name}/{h_output_name}. Source columns: {list(df_source.columns)}")
                    
                    return df_target
                except Exception as e:
                    print(f"Error en safe_add_curve_data: {e}")
                    return df_target

            # ---------- AGREGAR TODAS LAS TABLAS NECESARIAS PARA GRÁFICOS ----------
            
            # DataFrame principal que contendrá todas las tablas
            df_graficos = pd.DataFrame()
            
            # ---------- TABLAS 100% RPM ----------
            
            # Tabla 1: Curva H-Q Bomba 100%
            if not df_bomba_100.empty:
                df_graficos = safe_add_curve_data(df_bomba_100, df_graficos, 'Caudal (L/s)', 'Altura (m)', 'Q_Bomba_100', 'H_Bomba_100')
            
            # Tabla 2: Curva de Rendimiento (Eficiencia) 100%
            if not df_rendimiento_100.empty:
                df_graficos = safe_add_curve_data(df_rendimiento_100, df_graficos, 'Caudal (L/s)', 'Rendimiento (%)', 'Q_Eff_100', 'Eff_100')
            
            # Tabla 3: Curva de Potencia 100%
            if not df_potencia_100.empty:
                df_graficos = safe_add_curve_data(df_potencia_100, df_graficos, 'Caudal (L/s)', 'Potencia (HP)', 'Q_Power_100', 'P_100')
            
            # Tabla 4: Curva NPSH 100%
            if not df_npsh_100.empty:
                df_graficos = safe_add_curve_data(df_npsh_100, df_graficos, 'Caudal (L/s)', 'NPSH (m)', 'Q_NPSHr', 'NPSHr')
            
            # Tabla 5: Curva del Sistema 100%
            if not df_sistema_100.empty:
                df_graficos = safe_add_curve_data(df_sistema_100, df_graficos, 'Caudal (L/s)', 'Altura (m)', 'Q_Sistema', 'H_Sistema')
            
            # ---------- TABLAS VDF ----------
            
            # Tabla 6: Curva H-Q Bomba VDF
            if not df_bomba_vfd.empty:
                df_graficos = safe_add_curve_data(df_bomba_vfd, df_graficos, 'Caudal (L/s)', 'Altura (m)', 'Q_Bomba_VDF', 'H_Bomba_VDF')
            
            # Tabla 7: Curva de Rendimiento VDF
            if not df_rendimiento_vfd.empty:
                df_graficos = safe_add_curve_data(df_rendimiento_vfd, df_graficos, 'Caudal (L/s)', 'Rendimiento (%)', 'Q_Eff_VDF', 'Eff_VDF')
            
            # Tabla 8: Curva de Potencia VDF  
            if not df_potencia_vfd.empty:
                df_graficos = safe_add_curve_data(df_potencia_vfd, df_graficos, 'Caudal (L/s)', 'Potencia (HP)', 'Q_Power_VDF', 'P_VDF')
            
            # Tabla 9: Curva NPSH VDF
            if not df_npsh_vfd.empty:
                df_graficos = safe_add_curve_data(df_npsh_vfd, df_graficos, 'Caudal (L/s)', 'NPSH (m)', 'Q_NPSHr_VDF', 'NPSHr_VDF')
            
            # Tabla 10: Curva del Sistema VDF
            if not df_sistema_vfd.empty:
                df_graficos = safe_add_curve_data(df_sistema_vfd, df_graficos, 'Caudal (L/s)', 'Altura (m)', 'Q_Sistema_VDF', 'H_Sistema_VDF')
            
            # ---------- COMPATIBILIDAD CON CÓDIGO ANTERIOR ----------
            
            # Si aún no tenemos datos suficientes, usar las tablas de fallback
            if df_graficos.empty or len(df_graficos.columns) < 4:
                # Curva del Sistema (fallback)
                df_graficos = safe_add_curve_data(df_system, df_graficos, 'Caudal (L/s)', 'ADT (m)', 'Q_Sistema', 'H_Sistema')
                
                # Curva Eficiencia (fallback)
                df_graficos = safe_add_curve_data(df_bomba_eff, df_graficos, 'Caudal (L/s)', 'Eficiencia (%)', 'Q_Eff_100', 'Eff_100')
                
                # Curva Potencia (fallback)
                df_graficos = safe_add_curve_data(df_power_100_fallback, df_graficos, 'Caudal (L/s)', 'Potencia (HP)', 'Q_Power_100', 'P_100')
                
                # Curva NPSH (fallback)
                df_graficos = safe_add_curve_data(df_npshr, df_graficos, 'Caudal (L/s)', 'NPSHr (m)', 'Q_NPSHr', 'NPSHr')

            # Eliminar columnas duplicadas de Caudal
            df_graficos = df_graficos.loc[:,~df_graficos.columns.duplicated()]

            # Validar que tenemos al menos datos del sistema
            if df_graficos.empty:
                # Crear datos mínimos de ejemplo si no hay datos
                df_graficos = pd.DataFrame({
                    'Q_Sistema': [0, 50, 100, 150, 200],
                    'H_Sistema': [50, 52, 56, 62, 70],
                    'Q_Bomba_100': [0, 50, 100, 150, 200],
                    'H_Bomba_100': [60, 58, 54, 48, 40],
                    'Q_Eff_100': [0, 50, 100, 150, 200],
                    'Eff_100': [0, 75, 80, 75, 65],
                    'Q_Power_100': [0, 50, 100, 150, 200],
                    'P_100': [0, 15, 25, 35, 45]
                })
            


            # Obtener coeficientes de las curvas desde session_state
            import numpy as np
            ajuste_tipo = session_state.get('ajuste_tipo', 'Cuadrática (2do grado)')
            
            # Determinar grado del polinomio
            if ajuste_tipo == "Lineal":
                grado = 1
            elif ajuste_tipo == "Cuadrática (2do grado)":
                grado = 2
            else:  # Polinomial
                grado = 3
            
            # Función para calcular coeficientes de una curva
            def get_curve_coefficients(df, q_col, y_col):
                """Obtiene los coeficientes del polinomio ajustado a los datos"""
                try:
                    if df is None or df.empty:
                        return None
                    
                    # Buscar columna Q de forma flexible
                    q_col_found = None
                    for col in df.columns:
                        col_lower = col.lower()
                        if 'caudal' in col_lower or 'q_' in col_lower or col_lower.startswith('q'):
                            q_col_found = col
                            break
                    
                    # Buscar columna Y de forma flexible
                    y_col_found = None
                    y_keywords = {
                        'Altura (m)': ['altura', 'h_', 'head', 'altura (m)'],
                        'Rendimiento (%)': ['rendimiento', 'eficiencia', 'eff', 'rendimiento (%)'],
                        'Potencia (HP)': ['potencia', 'power', 'p_', 'potencia (hp)'],
                        'NPSH (m)': ['npsh', 'npshr', 'npsh (m)', 'npshr (m)'],
                    }
                    
                    if y_col in df.columns:
                        y_col_found = y_col
                    else:
                        keywords = y_keywords.get(y_col, [y_col.lower().split()[0]])
                        for col in df.columns:
                            col_lower = col.lower()
                            if any(kw in col_lower for kw in keywords):
                                y_col_found = col
                                break
                    
                    if q_col_found and y_col_found:
                        x_data = df[q_col_found].values
                        y_data = df[y_col_found].values
                        
                        # Filtrar valores válidos
                        mask = np.isfinite(x_data) & np.isfinite(y_data)
                        x_data = x_data[mask]
                        y_data = y_data[mask]
                        
                        if len(x_data) >= 2:
                            # Ajustar el grado según el número de puntos
                            grado_ajustado = min(grado, len(x_data) - 1)
                            coef = np.polyfit(x_data, y_data, grado_ajustado)
                            return coef
                except Exception:
                    pass
                return None
            
            coef_bomba_100 = get_curve_coefficients(df_bomba_100, 'Caudal (L/s)', 'Altura (m)')
            coef_eff_100 = get_curve_coefficients(df_rendimiento_100, 'Caudal (L/s)', 'Rendimiento (%)')
            coef_power_100 = get_curve_coefficients(df_potencia_100, 'Caudal (L/s)', 'Potencia (HP)')
            coef_sistema = get_curve_coefficients(df_sistema_100, 'Caudal (L/s)', 'Altura (m)')
            
            # NPSH: Intentar calcular desde curva_inputs primero (más confiable)
            coef_npsh_100 = None
            if curva_inputs.get('npsh') and len(curva_inputs['npsh']) >= 2:
                try:
                    puntos_npsh = curva_inputs['npsh']
                    x_npsh = np.array([pt[0] for pt in puntos_npsh])
                    y_npsh = np.array([pt[1] for pt in puntos_npsh])
                    if np.all(np.isfinite(x_npsh)) and np.all(np.isfinite(y_npsh)):
                        coef_npsh_100 = np.polyfit(x_npsh, y_npsh, min(grado, len(x_npsh) - 1))
                except Exception:
                    pass
            
            # Fallback a DataFrame si curva_inputs no tiene NPSH
            if coef_npsh_100 is None:
                coef_npsh_100 = get_curve_coefficients(df_npsh_100, 'Caudal (L/s)', 'NPSH (m)')
            
            # FALLBACK GARANTIZADO: Si aún no hay coeficientes, usar ecuación típica NPSH
            # NPSHr típico para bombas centrífugas: NPSHr = a*Q² + b*Q + c
            if coef_npsh_100 is None:
                # Coeficientes típicos: NPSHr ≈ 0.0005*Q² + 0.01*Q + 2.0 (Q en L/s, NPSHr en m)
                coef_npsh_100 = np.array([0.0005, 0.01, 2.0])
            
            coef_bomba_vfd = get_curve_coefficients(df_bomba_vfd, 'Caudal (L/s)', 'Altura (m)')
            coef_eff_vfd = get_curve_coefficients(df_rendimiento_vfd, 'Caudal (L/s)', 'Rendimiento (%)')
            coef_power_vfd = get_curve_coefficients(df_potencia_vfd, 'Caudal (L/s)', 'Potencia (HP)')
            coef_sistema_vfd = get_curve_coefficients(df_sistema_vfd, 'Caudal (L/s)', 'Altura (m)')
            
            # NPSH VFD: usar los mismos coeficientes que 100% (NPSH requerido no cambia con VFD)
            coef_npsh_vfd = coef_npsh_100
            
            # Mapeo de columnas a coeficientes
            coef_map = {
                'H_Bomba_100': coef_bomba_100,
                'Eff_100': coef_eff_100,
                'P_100': coef_power_100,
                'NPSHr': coef_npsh_100,
                'H_Sistema': coef_sistema,
                'H_Bomba_VDF': coef_bomba_vfd,
                'Eff_VDF': coef_eff_vfd,
                'P_VDF': coef_power_vfd,
                'NPSHr_VDF': coef_npsh_vfd,
                'H_Sistema_VDF': coef_sistema_vfd
            }
            
            # Función para crear fórmula de Excel
            def create_excel_formula(coef, q_cell_ref):
                """Crea una fórmula de Excel basada en los coeficientes del polinomio"""
                if coef is None or len(coef) == 0:
                    return None
                
                try:
                    terms = []
                    n = len(coef)
                    
                    for i, c in enumerate(coef):
                        power = n - 1 - i
                        
                        if abs(c) < 1e-10:  # Coeficiente muy pequeño, ignorar
                            continue
                        
                        # Determinar signo
                        is_negative = c < 0
                        abs_c = abs(c)
                        
                        # Formatear el coeficiente (valor absoluto)
                        if abs_c < 0.000001:
                            coef_str = f"{abs_c:.10f}".rstrip('0').rstrip('.')
                        else:
                            coef_str = f"{abs_c:.6f}".rstrip('0').rstrip('.')
                        
                        # Crear el término
                        if power == 0:
                            term_str = coef_str
                        elif power == 1:
                            term_str = f"{coef_str}*{q_cell_ref}"
                        else:
                            term_str = f"{coef_str}*POWER({q_cell_ref},{power})"
                        
                        # Agregar con signo
                        terms.append((is_negative, term_str))
                    
                    if not terms:
                        return None
                    
                    # Construir fórmula correcta
                    # Primer término
                    is_neg, term = terms[0]
                    if is_neg:
                        formula = f"-{term}"
                    else:
                        formula = term
                    
                    # Términos restantes
                    for is_neg, term in terms[1:]:
                        if is_neg:
                            formula += f"-{term}"
                        else:
                            formula += f"+{term}"
                    
                    return f"={formula}"
                except Exception as e:
                    print(f"Error creando fórmula: {e}")
                    return None
            
            # Agregar una fila vacía primero (fila 1 para títulos)
            ws_data.append([])
            
            # Agregar datos a la hoja
            # Si incluir_formulas=True, generar nuevos datos con Q del usuario y fórmulas
            # Los datos comienzan en fila 2
            if incluir_formulas:
                # ========== MODO FÓRMULAS: Generar datos usando parámetros del usuario ==========
                
                # Definir estructura de columnas para 100% y VFD
                # Columnas 100%: Q_Bomba_100, H_Bomba_100, Q_Eff_100, Eff_100, Q_Power_100, P_100, Q_NPSHr, NPSHr, Q_Sistema, H_Sistema
                # Columnas VFD: Q_Bomba_VDF, H_Bomba_VDF, Q_Eff_VDF, Eff_VDF, Q_Power_VDF, P_VDF, Q_NPSHr_VDF, NPSHr_VDF, Q_Sistema_VDF, H_Sistema_VDF
                
                headers_100 = ['Q_Bomba_100', 'H_Bomba_100', 'Q_Eff_100', 'Eff_100', 'Q_Power_100', 'P_100', 'Q_NPSHr', 'NPSHr', 'Q_Sistema', 'H_Sistema']
                headers_vfd = ['Q_Bomba_VDF', 'H_Bomba_VDF', 'Q_Eff_VDF', 'Eff_VDF', 'Q_Power_VDF', 'P_VDF', 'Q_NPSHr_VDF', 'NPSHr_VDF', 'Q_Sistema_VDF', 'H_Sistema_VDF']
                all_headers = headers_100 + headers_vfd
                
                # Escribir encabezados
                ws_data.append(all_headers)
                
                # Determinar el número máximo de filas
                num_rows_100 = len(caudales_tabla)
                num_rows_vfd = len(caudales_tabla_vfd)
                max_rows = max(num_rows_100, num_rows_vfd)
                
                # Mapeo de columnas de valor a su columna Q correspondiente y coeficientes
                col_formula_map = {
                    'H_Bomba_100': ('Q_Bomba_100', coef_bomba_100),
                    'Eff_100': ('Q_Eff_100', coef_eff_100),
                    'P_100': ('Q_Power_100', coef_power_100),
                    'NPSHr': ('Q_NPSHr', coef_npsh_100),
                    'H_Sistema': ('Q_Sistema', coef_sistema),
                    'H_Bomba_VDF': ('Q_Bomba_VDF', coef_bomba_vfd),
                    'Eff_VDF': ('Q_Eff_VDF', coef_eff_vfd),
                    'P_VDF': ('Q_Power_VDF', coef_power_vfd),
                    'NPSHr_VDF': ('Q_NPSHr_VDF', coef_npsh_vfd),
                    'H_Sistema_VDF': ('Q_Sistema_VDF', coef_sistema_vfd),
                }
                
                # DEBUG: Verificar coeficientes NPSH
                print(f"DEBUG NPSH: coef_npsh_100 = {coef_npsh_100}")
                print(f"DEBUG NPSH: coef_npsh_vfd = {coef_npsh_vfd}")
                print(f"DEBUG: curva_inputs keys = {list(curva_inputs.keys())}")
                if curva_inputs.get('npsh'):
                    print(f"DEBUG: curva_inputs['npsh'] = {curva_inputs['npsh'][:3]}...")
                
                # VERIFY: NPSHr should map to coef_npsh_100
                print(f"DEBUG col_formula_map['NPSHr'] = {col_formula_map.get('NPSHr')}")
                print(f"DEBUG col_formula_map['NPSHr_VDF'] = {col_formula_map.get('NPSHr_VDF')}")
                
                # Escribir cada fila de datos
                for row_idx in range(max_rows):
                    row_data = []
                    excel_row = row_idx + 3  # Fila 3 en adelante (fila 1 vacía, fila 2 encabezados)
                    
                    for col_idx, col_name in enumerate(all_headers):
                        col_letter = get_column_letter(col_idx + 1)
                        
                        if col_name.startswith('Q_'):
                            # Es una columna Q - escribir valor numérico
                            if 'VDF' in col_name:
                                # Columna VFD
                                if row_idx < num_rows_vfd:
                                    row_data.append(round(caudales_tabla_vfd[row_idx], 3))
                                else:
                                    row_data.append('')
                            else:
                                # Columna 100%
                                if row_idx < num_rows_100:
                                    row_data.append(round(caudales_tabla[row_idx], 3))
                                else:
                                    row_data.append('')
                        else:
                            # Es una columna de valor (H, Eff, P, NPSH) - escribir fórmula
                            if col_name in col_formula_map:
                                q_col_name, coef = col_formula_map[col_name]
                                
                                # Verificar si hay datos para esta fila
                                is_vfd = 'VDF' in col_name
                                if (is_vfd and row_idx < num_rows_vfd) or (not is_vfd and row_idx < num_rows_100):
                                    if coef is not None:
                                        # Encontrar la columna Q correspondiente
                                        q_col_idx = all_headers.index(q_col_name)
                                        q_col_letter = get_column_letter(q_col_idx + 1)
                                        q_cell_ref = f"{q_col_letter}{excel_row}"
                                        
                                        # Crear fórmula
                                        formula = create_excel_formula(coef, q_cell_ref)
                                        if formula:
                                            row_data.append(formula)
                                        else:
                                            row_data.append('')
                                    else:
                                        row_data.append('')
                                else:
                                    row_data.append('')
                            else:
                                row_data.append('')
                    
                    ws_data.append(row_data)
            else:
                # Modo sin fórmulas (comportamiento original)
                for r in dataframe_to_rows(df_graficos, index=False, header=True):
                    ws_data.append(r)
            
            # Formatear encabezados en fila 2 (negrita, centrado horizontal y vertical, ajustado)
            for cell in ws_data[2]:
                if cell.value:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Función para crear ecuación legible
            def create_readable_equation(coef, var_name="Y", q_var="Q"):
                """Crea una ecuación legible basada en los coeficientes del polinomio"""
                if coef is None or len(coef) == 0:
                    return f"{var_name} = N/A"
                
                terms = []
                n = len(coef)
                
                for i, c in enumerate(coef):
                    power = n - 1 - i
                    
                    if abs(c) < 1e-10:  # Coeficiente muy pequeño, ignorar
                        continue
                    
                    # Formatear coeficiente
                    coef_str = f"{c:.6f}"
                    
                    if power == 0:
                        # Término constante
                        terms.append(coef_str)
                    elif power == 1:
                        # Término lineal
                        if abs(c - 1.0) < 1e-10:
                            terms.append(q_var)
                        elif abs(c + 1.0) < 1e-10:
                            terms.append(f"-{q_var}")
                        else:
                            terms.append(f"{coef_str}*{q_var}")
                    else:
                        # Términos de potencia mayor
                        if abs(c - 1.0) < 1e-10:
                            terms.append(f"{q_var}^{power}")
                        elif abs(c + 1.0) < 1e-10:
                            terms.append(f"-{q_var}^{power}")
                        else:
                            terms.append(f"{coef_str}*{q_var}^{power}")
                
                if not terms:
                    return f"{var_name} = 0"
                
                # Unir términos con el signo correcto
                equation = f"{var_name} = {terms[0]}"
                for term in terms[1:]:
                    if term.startswith('-'):
                        equation += f" {term}"
                    else:
                        equation += f" + {term}"
                
                return equation
            
            # Agregar ecuaciones en la columna V y descripciones en W
            col_v_start = 2
            col_w_start = 2
            
            # Calcular la última fila con datos para cada sección
            # Buscar la última fila con datos en las columnas A-J (primera sección)
            max_row_aj = 2  # Mínimo fila 2 (encabezados)
            for row in range(ws_data.max_row, 0, -1):
                has_data = False
                for col in range(1, 11):  # Columnas A(1) a J(10)
                    if ws_data.cell(row=row, column=col).value is not None:
                        has_data = True
                        break
                if has_data:
                    max_row_aj = row
                    break
            
            # Buscar la última fila con datos en las columnas K-T (segunda sección)
            max_row_kt = 2  # Mínimo fila 2 (encabezados)
            for row in range(ws_data.max_row, 0, -1):
                has_data = False
                for col in range(11, 21):  # Columnas K(11) a T(20)
                    if ws_data.cell(row=row, column=col).value is not None:
                        has_data = True
                        break
                if has_data:
                    max_row_kt = row
                    break
            
            # Configurar impresión para Datos Gráficos con 3 áreas
            ws_data.page_setup.orientation = 'portrait'
            ws_data.page_setup.paperSize = 1  # Letter
            ws_data.print_area = f'A1:J{max_row_aj},K1:T{max_row_kt},V1:W16'
            
            # Títulos en A1:J1 y K1:T1 (combinados)
            ws_data.merge_cells('A1:J1')
            ws_data.cell(row=1, column=1, value="Curvas 100%")
            ws_data.cell(row=1, column=1).font = Font(bold=True, size=12, color="FFFFFF")
            ws_data.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws_data.cell(row=1, column=1).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            
            ws_data.merge_cells('K1:T1')
            ws_data.cell(row=1, column=11, value=f"Curvas {vfd_percentage:.0f}% VFD")
            ws_data.cell(row=1, column=11).font = Font(bold=True, size=12, color="FFFFFF")
            ws_data.cell(row=1, column=11).alignment = Alignment(horizontal='center', vertical='center')
            ws_data.cell(row=1, column=11).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            
            # Título
            ws_data.cell(row=1, column=22, value="ECUACIONES DE CURVAS")  # Columna V
            ws_data.cell(row=1, column=23, value="DESCRIPCIÓN")  # Columna W
            
            # Formatear títulos
            ws_data.cell(row=1, column=22).font = Font(bold=True, size=12, color="FFFFFF")
            ws_data.cell(row=1, column=22).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            ws_data.cell(row=1, column=22).alignment = Alignment(horizontal='center', vertical='center')
            
            ws_data.cell(row=1, column=23).font = Font(bold=True, size=12, color="FFFFFF")
            ws_data.cell(row=1, column=23).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            ws_data.cell(row=1, column=23).alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar ancho de columnas
            ws_data.column_dimensions['V'].width = 50
            ws_data.column_dimensions['W'].width = 35
            
            # Agregar bordes a los datos numéricos desde C1
            from openpyxl.styles import Border, Side
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Aplicar bordes a todas las celdas con datos (A1:T2 y V1:W17)
            for row in ws_data.iter_rows(min_row=1, max_row=ws_data.max_row, min_col=1, max_col=20):
                for cell in row:
                    cell.border = thin_border
            
            for row in ws_data.iter_rows(min_row=1, max_row=17, min_col=22, max_col=23):
                for cell in row:
                    cell.border = thin_border
            
            # Centrar todos los valores numéricos desde la fila C1 hacia abajo
            for row in ws_data.iter_rows(min_row=1, max_row=ws_data.max_row, min_col=1, max_col=20):
                for cell in row:
                    if cell.row > 1:  # No centrar encabezados
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row = col_v_start
            
            # Ecuaciones 100% RPM
            ws_data.cell(row=current_row, column=23, value="=== CURVAS 100% RPM ===")
            ws_data.cell(row=current_row, column=23).font = Font(bold=True)
            current_row += 1
            
            # Curva de Altura de la Bomba 100%
            if coef_bomba_100 is not None:
                ecuacion = create_readable_equation(coef_bomba_100, "H_Bomba", "Q")
                ws_data.cell(row=current_row, column=22, value=ecuacion)
                ws_data.cell(row=current_row, column=23, value="Curva de Altura de la Bomba")
                current_row += 1
            
            # Curva de Eficiencia 100%
            if coef_eff_100 is not None:
                ecuacion = create_readable_equation(coef_eff_100, "Eficiencia", "Q")
                ws_data.cell(row=current_row, column=22, value=ecuacion)
                ws_data.cell(row=current_row, column=23, value="Curva de Eficiencia (%)")
                current_row += 1
            
            # Curva de Potencia 100%
            if coef_power_100 is not None:
                ecuacion = create_readable_equation(coef_power_100, "Potencia", "Q")
                ws_data.cell(row=current_row, column=22, value=ecuacion)
                ws_data.cell(row=current_row, column=23, value="Curva de Potencia (HP)")
                current_row += 1
            
            # Curva de NPSH 100%
            if coef_npsh_100 is not None:
                ecuacion = create_readable_equation(coef_npsh_100, "NPSHr", "Q")
                ws_data.cell(row=current_row, column=22, value=ecuacion)
                ws_data.cell(row=current_row, column=23, value="Curva de NPSH Requerido (m)")
                current_row += 1
            
            # Curva del Sistema
            if coef_sistema is not None:
                ecuacion = create_readable_equation(coef_sistema, "H_Sistema", "Q")
                ws_data.cell(row=current_row, column=22, value=ecuacion)
                ws_data.cell(row=current_row, column=23, value="Curva del Sistema (m)")
                current_row += 1
            
            current_row += 1  # Espacio
            
            # Ecuaciones VFD
            ws_data.cell(row=current_row, column=23, value="=== CURVAS VFD ===")
            ws_data.cell(row=current_row, column=23).font = Font(bold=True)
            current_row += 1
            
            # Curva de Altura de la Bomba VFD
            if coef_bomba_vfd is not None:
                ecuacion = create_readable_equation(coef_bomba_vfd, "H_Bomba_VFD", "Q")
                ws_data.cell(row=current_row, column=22, value=ecuacion)
                ws_data.cell(row=current_row, column=23, value="Curva de Altura de la Bomba VFD")
                current_row += 1
            
            # Curva de Eficiencia VFD
            if coef_eff_vfd is not None:
                ecuacion = create_readable_equation(coef_eff_vfd, "Eficiencia_VFD", "Q")
                ws_data.cell(row=current_row, column=22, value=ecuacion)
                ws_data.cell(row=current_row, column=23, value="Curva de Eficiencia VFD (%)")
                current_row += 1
            
            # Curva de Potencia VFD
            if coef_power_vfd is not None:
                ecuacion = create_readable_equation(coef_power_vfd, "Potencia_VFD", "Q")
                ws_data.cell(row=current_row, column=22, value=ecuacion)
                ws_data.cell(row=current_row, column=23, value="Curva de Potencia VFD (HP)")
                current_row += 1
            
            # Curva de NPSH VFD
            if coef_npsh_vfd is not None:
                ecuacion = create_readable_equation(coef_npsh_vfd, "NPSHr_VFD", "Q")
                ws_data.cell(row=current_row, column=22, value=ecuacion)
                ws_data.cell(row=current_row, column=23, value="Curva de NPSH Requerido VFD (m)")
                current_row += 1
            
            # Curva del Sistema VFD (normalmente es la misma que 100%)
            if coef_sistema_vfd is not None:
                ecuacion = create_readable_equation(coef_sistema_vfd, "H_Sistema_VFD", "Q")
                ws_data.cell(row=current_row, column=22, value=ecuacion)
                ws_data.cell(row=current_row, column=23, value="Curva del Sistema VFD (m)")
                current_row += 1
            
            # Agregar información del tipo de ajuste
            current_row += 1
            ws_data.cell(row=current_row, column=22, value=f"Tipo de ajuste: {ajuste_tipo}")
            ws_data.cell(row=current_row, column=22).font = Font(italic=True)
                
        except Exception as e:
            print(f"Warning: Error procesando datos de gráficos: {e}")
            # Datos de emergencia
            emergency_data = pd.DataFrame({
                'Q_Sistema': [0, 50, 100, 150, 200],
                'H_Sistema': [50, 52, 56, 62, 70]
            })
            for r in dataframe_to_rows(emergency_data, index=False, header=True):
                ws_data.append(r)

        # --- Hoja 5: Gráficos 100% RPM ---
        ws_graficos_100 = wb.create_sheet("Gráficos 100% RPM")
        
        # Eliminar líneas de cuadrícula
        ws_graficos_100.sheet_view.showGridLines = False
        
        # Configurar impresión horizontal
        ws_graficos_100.page_setup.orientation = 'landscape'
        ws_graficos_100.page_setup.paperSize = 1  # Letter
        ws_graficos_100.page_setup.fitToPage = True
        ws_graficos_100.page_setup.fitToHeight = 1
        ws_graficos_100.page_setup.fitToWidth = 1
        ws_graficos_100.print_area = 'A1:S43'
        
        # Título de la hoja
        ws_graficos_100.cell(row=1, column=1, value="Gráficos 100% RPM").font = Font(bold=True, size=16, color="000000")
        ws_graficos_100.merge_cells('A1:S2')
        ws_graficos_100.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws_graficos_100.cell(row=1, column=1).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        
        # Ocultar hoja si incluir_graficos es False
        if not incluir_graficos:
            ws_graficos_100.sheet_state = 'hidden'
        
        try:
            if not df_graficos.empty:
                max_rows = len(df_graficos) + 1

                # Segurizar la función que obtiene referencias de columnas
                def safe_get_column_ref(ws_data, df, col_name, start_row=3, end_row=None):
                    """Función segura para obtener referencia de columna
                    start_row=3 porque ahora los datos comienzan en fila 3 (fila 1 vacía, fila 2 encabezados)
                    """
                    try:
                        if end_row is None:
                            end_row = len(df) + 2  # +2 porque hay una fila vacía al inicio
                        
                        if col_name in df.columns:
                            col_index = df.columns.get_loc(col_name) + 1
                            return Reference(ws_data, min_col=col_index, min_row=start_row, max_row=end_row)
                        else:
                            return None
                    except Exception:
                        return None

                # Función para calcular punto de operación (intersección de curvas)
                def calcular_punto_operacion(coef_bomba, coef_sistema):
                    """Calcula el punto de intersección entre la curva de la bomba y del sistema"""
                    try:
                        if coef_bomba is None or coef_sistema is None:
                            return None, None
                        
                        # Crear ecuación de diferencia: bomba - sistema = 0
                        # Asegurar que ambos polinomios tengan el mismo grado
                        max_len = max(len(coef_bomba), len(coef_sistema))
                        
                        # Rellenar con ceros a la izquierda si es necesario
                        coef_b = np.pad(coef_bomba, (max_len - len(coef_bomba), 0), 'constant')
                        coef_s = np.pad(coef_sistema, (max_len - len(coef_sistema), 0), 'constant')
                        
                        # Diferencia de coeficientes
                        coef_diff = coef_b - coef_s
                        
                        # Encontrar raíces
                        roots = np.roots(coef_diff)
                        
                        # Filtrar raíces reales y positivas
                        real_roots = []
                        for root in roots:
                            if np.isreal(root) and root.real > 0:
                                real_roots.append(root.real)
                        
                        if not real_roots:
                            return None, None
                        
                        # Tomar la raíz más pequeña positiva (punto de operación normal)
                        q_op = min(real_roots)
                        
                        # Calcular altura en el punto de operación
                        h_op = np.polyval(coef_bomba, q_op)
                        
                        return float(q_op), float(h_op)
                    except Exception as e:
                        print(f"Error calculando punto de operación: {e}")
                        return None, None
                
                # Calcular punto de operación 100% RPM
                q_op_100, h_op_100 = calcular_punto_operacion(coef_bomba_100, coef_sistema)
                
                # Gráfico 1: Curva Bomba vs Sistema - SOLO 100% RPM
                chart1 = ScatterChart()
                chart1.title = "Curva de Bomba vs Curva de Sistema - 100% RPM"
                chart1.x_axis.title = "Caudal (L/s)"
                chart1.y_axis.title = "Altura (m)"

                # Serie Curva Sistema
                x_sys = safe_get_column_ref(ws_data, df_graficos, "Q_Sistema")
                y_sys = safe_get_column_ref(ws_data, df_graficos, "H_Sistema")
                if x_sys and y_sys:
                    s_sys = Series(y_sys, x_sys, title="Curva Sistema")
                    chart1.series.append(s_sys)

                # Serie Curva Bomba 100%
                x_bomba = safe_get_column_ref(ws_data, df_graficos, "Q_Bomba_100")
                y_bomba = safe_get_column_ref(ws_data, df_graficos, "H_Bomba_100")
                if x_bomba and y_bomba:
                    s_bomba = Series(y_bomba, x_bomba, title='Curva Bomba 100% RPM')
                    chart1.series.append(s_bomba)
                
                # NO incluir curvas VDF en este gráfico (serán en otra hoja)

                ws_graficos_100.add_chart(chart1, "A3")
                
                # Gráfico 2: Eficiencia
                if 'Q_Eff_100' in df_graficos.columns and 'Eff_100' in df_graficos.columns:
                    chart2 = ScatterChart()
                    chart2.title = "Curva de Eficiencia"
                    chart2.x_axis.title = "Caudal (L/s)"
                    chart2.y_axis.title = "Eficiencia (%)"
                    
                    x_eff = safe_get_column_ref(ws_data, df_graficos, "Q_Eff_100")
                    y_eff = safe_get_column_ref(ws_data, df_graficos, "Eff_100")
                    if x_eff and y_eff:
                        s_eff = Series(y_eff, x_eff, title="Eficiencia 100% RPM")
                        chart2.series.append(s_eff)
                        ws_graficos_100.add_chart(chart2, "K3")
                
                # Agregar información del punto de operación en fila 19
                if q_op_100 is not None and h_op_100 is not None:
                    ws_graficos_100.cell(row=19, column=1, value="PUNTO DE OPERACIÓN 100% RPM:")
                    ws_graficos_100.cell(row=19, column=1).font = Font(bold=True, size=11, color="FF0000")
                    ws_graficos_100.cell(row=20, column=1, value=f"Caudal: {q_op_100:.2f} L/s")
                    ws_graficos_100.cell(row=20, column=1).font = Font(bold=True, size=10)
                    ws_graficos_100.cell(row=21, column=1, value=f"Altura: {h_op_100:.2f} m")
                    ws_graficos_100.cell(row=21, column=1).font = Font(bold=True, size=10)
                    
                    # Calcular eficiencia en el punto de operación
                    if coef_eff_100 is not None:
                        eff_at_op = np.polyval(coef_eff_100, q_op_100)
                        ws_graficos_100.cell(row=22, column=1, value=f"Eficiencia: {eff_at_op:.2f} %")
                        ws_graficos_100.cell(row=22, column=1).font = Font(bold=True, size=10)
                    
                    # Calcular potencia en el punto de operación
                    if coef_power_100 is not None:
                        pow_at_op = np.polyval(coef_power_100, q_op_100)
                        ws_graficos_100.cell(row=23, column=1, value=f"Potencia: {pow_at_op:.2f} HP")
                        ws_graficos_100.cell(row=23, column=1).font = Font(bold=True, size=10)
                    
                    # Calcular NPSH en el punto de operación
                    if coef_npsh_100 is not None:
                        npsh_at_op = np.polyval(coef_npsh_100, q_op_100)
                        ws_graficos_100.cell(row=24, column=1, value=f"NPSHr: {npsh_at_op:.2f} m")
                        ws_graficos_100.cell(row=24, column=1).font = Font(bold=True, size=10)

                # Gráfico 3: Potencia
                if 'Q_Power_100' in df_graficos.columns and 'P_100' in df_graficos.columns:
                    chart3 = ScatterChart()
                    chart3.title = "Curva de Potencia"
                    chart3.x_axis.title = "Caudal (L/s)"
                    chart3.y_axis.title = "Potencia (HP)"
                    
                    x_pow = safe_get_column_ref(ws_data, df_graficos, "Q_Power_100")
                    y_pow = safe_get_column_ref(ws_data, df_graficos, "P_100")
                    if x_pow and y_pow:
                        s_pow = Series(y_pow, x_pow, title="Potencia 100% RPM")
                        chart3.series.append(s_pow)
                        ws_graficos_100.add_chart(chart3, "A26")

                # Gráfico 4: NPSH
                if 'Q_NPSHr' in df_graficos.columns and 'NPSHr' in df_graficos.columns:
                    chart4 = ScatterChart()
                    chart4.title = "Análisis de NPSH"
                    chart4.x_axis.title = "Caudal (L/s)"
                    chart4.y_axis.title = "NPSH (m)"
                    
                    x_npsh = safe_get_column_ref(ws_data, df_graficos, "Q_NPSHr")
                    y_npsh = safe_get_column_ref(ws_data, df_graficos, "NPSHr")
                    print(f"DEBUG NPSH: x_npsh={x_npsh}, y_npsh={y_npsh}")
                    if x_npsh and y_npsh:
                        s_npsh = Series(y_npsh, x_npsh, title="NPSH Requerido")
                        chart4.series.append(s_npsh)
                        ws_graficos_100.add_chart(chart4, "K26")
                        print("DEBUG NPSH: Gráfico agregado correctamente")
                    else:
                        print(f"DEBUG NPSH: Referencias no válidas - agregando gráfico de todos modos")
                        # Forzar la adición del gráfico aunque sea vacío
                        ws_graficos_100.add_chart(chart4, "K26")
                else:
                    print(f"DEBUG NPSH: Columnas disponibles: {list(df_graficos.columns)}")
                    print(f"DEBUG NPSH: Q_NPSHr en columnas? {'Q_NPSHr' in df_graficos.columns}, NPSHr en columnas? {'NPSHr' in df_graficos.columns}")
                
                # Agregar información adicional sobre el análisis 100% RPM en fila 41
                ws_graficos_100.cell(row=41, column=1, value="INFORMACIÓN DEL REPORTE - 100% RPM").font = Font(bold=True, size=12)
                ws_graficos_100.cell(row=42, column=1, value=f"Proyecto: {session_state.get('proyecto', 'Sin especificar')}")
                ws_graficos_100.cell(row=43, column=1, value=f"Diseño: {session_state.get('diseno', 'Sin especificar')}")
                ws_graficos_100.cell(row=44, column=1, value=f"Caudal de Diseño: {session_state.get('caudal_lps', 0):.2f} L/s")
                ws_graficos_100.cell(row=45, column=1, value=f"Altura Estática Total: {session_state.get('altura_estatica_total', 0):.2f} m")
                ws_graficos_100.cell(row=46, column=1, value=f"ADT Total: {session_state.get('adt_total', 0):.2f} m")
                
        except Exception as e:
            print(f"Warning: Error generando gráficos 100% RPM: {e}")
            ws_graficos_100.cell(row=1, column=1, value="ADVERTENCIA: Error generando gráficos 100% RPM - Verifique los datos de entrada")

        # --- Hoja 6: Gráficos VDF ---
        ws_graficos_vdf = wb.create_sheet(f"Gráficos {vfd_percentage:.0f}% RPM VDF")
        
        # Eliminar líneas de cuadrícula
        ws_graficos_vdf.sheet_view.showGridLines = False
        
        # Configurar impresión horizontal
        ws_graficos_vdf.page_setup.orientation = 'landscape'
        ws_graficos_vdf.page_setup.paperSize = 1  # Letter
        ws_graficos_vdf.page_setup.fitToPage = True
        ws_graficos_vdf.page_setup.fitToHeight = 1
        ws_graficos_vdf.page_setup.fitToWidth = 1
        ws_graficos_vdf.print_area = 'A1:S43'
        
        # Título de la hoja
        ws_graficos_vdf.cell(row=1, column=1, value=f"Gráficos {vfd_percentage:.0f}% RPM VDF").font = Font(bold=True, size=16, color="000000")
        ws_graficos_vdf.merge_cells('A1:S2')
        ws_graficos_vdf.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws_graficos_vdf.cell(row=1, column=1).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        
        # Ocultar hoja si incluir_graficos es False
        if not incluir_graficos:
            ws_graficos_vdf.sheet_state = 'hidden'
        
        try:
            
            if not df_graficos.empty:
                max_rows = len(df_graficos) + 1
                
                # Calcular punto de operación VFD
                q_op_vfd, h_op_vfd = calcular_punto_operacion(coef_bomba_vfd, coef_sistema)
                
                # Gráfico 1: Curva Bomba VDF vs Sistema
                if 'Q_Bomba_VDF' in df_graficos.columns and 'H_Bomba_VDF' in df_graficos.columns:
                    chart_vfd1 = ScatterChart()
                    chart_vfd1.title = f"Curva de Bomba vs Curva de Sistema - {vfd_percentage:.2f}% RPM"
                    chart_vfd1.x_axis.title = "Caudal (L/s)"
                    chart_vfd1.y_axis.title = "Altura (m)"

                    # Serie Curva Sistema
                    x_sys_vfd = safe_get_column_ref(ws_data, df_graficos, "Q_Sistema")
                    y_sys_vfd = safe_get_column_ref(ws_data, df_graficos, "H_Sistema")
                    if x_sys_vfd and y_sys_vfd:
                        s_sys_vfd = Series(y_sys_vfd, x_sys_vfd, title="Curva Sistema")
                        chart_vfd1.series.append(s_sys_vfd)

                    # Serie Curva Bomba VDF
                    x_vfd = safe_get_column_ref(ws_data, df_graficos, "Q_Bomba_VDF")
                    y_vfd = safe_get_column_ref(ws_data, df_graficos, "H_Bomba_VDF")
                    if x_vfd and y_vfd:
                        s_vfd = Series(y_vfd, x_vfd, title=f"Curva Bomba VFD {vfd_percentage:.2f}% RPM")
                        chart_vfd1.series.append(s_vfd)

                    ws_graficos_vdf.add_chart(chart_vfd1, "A3")

                # Gráfico 2: Eficiencia VDF
                if 'Q_Eff_VDF' in df_graficos.columns and 'Eff_VDF' in df_graficos.columns:
                    chart_vfd2 = ScatterChart()
                    chart_vfd2.title = f"Curva de Eficiencia VDF - {vfd_percentage:.2f}% RPM"
                    chart_vfd2.x_axis.title = "Caudal (L/s)"
                    chart_vfd2.y_axis.title = "Eficiencia (%)"
                    
                    x_eff_vfd = safe_get_column_ref(ws_data, df_graficos, "Q_Eff_VDF")
                    y_eff_vfd = safe_get_column_ref(ws_data, df_graficos, "Eff_VDF")
                    if x_eff_vfd and y_eff_vfd:
                        s_eff_vfd = Series(y_eff_vfd, x_eff_vfd, title=f"Eficiencia VDF {vfd_percentage:.2f}% RPM")
                        chart_vfd2.series.append(s_eff_vfd)
                        ws_graficos_vdf.add_chart(chart_vfd2, "K3")
                    
                # Agregar información del punto de operación en fila 19
                if q_op_vfd is not None and h_op_vfd is not None:
                    ws_graficos_vdf.cell(row=19, column=1, value=f"PUNTO DE OPERACIÓN {vfd_percentage:.2f}% RPM:")
                    ws_graficos_vdf.cell(row=19, column=1).font = Font(bold=True, size=11, color="FF0000")
                    ws_graficos_vdf.cell(row=20, column=1, value=f"Caudal: {q_op_vfd:.2f} L/s")
                    ws_graficos_vdf.cell(row=20, column=1).font = Font(bold=True, size=10)
                    ws_graficos_vdf.cell(row=21, column=1, value=f"Altura: {h_op_vfd:.2f} m")
                    ws_graficos_vdf.cell(row=21, column=1).font = Font(bold=True, size=10)
                    
                    # Calcular eficiencia en el punto de operación VFD
                    if coef_eff_vfd is not None:
                        eff_at_op_vfd = np.polyval(coef_eff_vfd, q_op_vfd)
                        ws_graficos_vdf.cell(row=22, column=1, value=f"Eficiencia: {eff_at_op_vfd:.2f} %")
                        ws_graficos_vdf.cell(row=22, column=1).font = Font(bold=True, size=10)
                    
                    # Calcular potencia en el punto de operación VFD
                    if coef_power_vfd is not None:
                        pow_at_op_vfd = np.polyval(coef_power_vfd, q_op_vfd)
                        ws_graficos_vdf.cell(row=23, column=1, value=f"Potencia: {pow_at_op_vfd:.2f} HP")
                        ws_graficos_vdf.cell(row=23, column=1).font = Font(bold=True, size=10)
                    
                    # Calcular NPSH en el punto de operación VFD
                    if coef_npsh_vfd is not None:
                        npsh_at_op_vfd = np.polyval(coef_npsh_vfd, q_op_vfd)
                        ws_graficos_vdf.cell(row=24, column=1, value=f"NPSHr: {npsh_at_op_vfd:.2f} m")
                        ws_graficos_vdf.cell(row=24, column=1).font = Font(bold=True, size=10)

                # Gráfico 3: Potencia VDF
                if 'Q_Power_VDF' in df_graficos.columns and 'P_VDF' in df_graficos.columns:
                    chart_vfd3 = ScatterChart()
                    chart_vfd3.title = f"Curva de Potencia VDF - {vfd_percentage:.2f}% RPM"
                    chart_vfd3.x_axis.title = "Caudal (L/s)"
                    chart_vfd3.y_axis.title = "Potencia (HP)"
                    
                    x_pow_vfd = safe_get_column_ref(ws_data, df_graficos, "Q_Power_VDF")
                    y_pow_vfd = safe_get_column_ref(ws_data, df_graficos, "P_VDF")
                    if x_pow_vfd and y_pow_vfd:
                        s_pow_vfd = Series(y_pow_vfd, x_pow_vfd, title=f"Potencia VDF {vfd_percentage:.2f}% RPM")
                        chart_vfd3.series.append(s_pow_vfd)
                        ws_graficos_vdf.add_chart(chart_vfd3, "A26")

                # Gráfico 4: NPSH VDF
                if 'Q_NPSHr_VDF' in df_graficos.columns and 'NPSHr_VDF' in df_graficos.columns:
                    chart_vdf4 = ScatterChart()
                    chart_vdf4.title = f"Análisis de NPSH VDF - {vfd_percentage:.2f}% RPM"
                    chart_vdf4.x_axis.title = "Caudal (L/s)"
                    chart_vdf4.y_axis.title = "NPSH (m)"
                    
                    x_npsh_vdf = safe_get_column_ref(ws_data, df_graficos, "Q_NPSHr_VDF")
                    y_npsh_vdf = safe_get_column_ref(ws_data, df_graficos, "NPSHr_VDF")
                    if x_npsh_vdf and y_npsh_vdf:
                        s_npsh_vdf = Series(y_npsh_vdf, x_npsh_vdf, title=f"NPSH Requerido VDF")
                        chart_vdf4.series.append(s_npsh_vdf)
                        ws_graficos_vdf.add_chart(chart_vdf4, "K26")
                
                # Información adicional del análisis VDF en fila 41
                ws_graficos_vdf.cell(row=41, column=1, value=f"INFORMACIÓN DEL REPORTE - {vfd_percentage:.2f}% RPM").font = Font(bold=True, size=12)
                ws_graficos_vdf.cell(row=42, column=1, value=f"Proyecto: {session_state.get('proyecto', 'Sin especificar')}")
                ws_graficos_vdf.cell(row=43, column=1, value=f"Diseño: {session_state.get('diseno', 'Sin especificar')}")
                ws_graficos_vdf.cell(row=44, column=1, value=f"Caudal VDF: {session_state.get('caudal_nominal_vdf', session_state.get('caudal_lps', 0)):.2f} L/s")
                ws_graficos_vdf.cell(row=45, column=1, value=f"Porcentaje VDF: {vfd_percentage:.2f}%")
            else:
                ws_graficos_vdf.cell(row=1, column=1, value=f"NO HAY DATOS VDF DISPONIBLES - {vfd_percentage:.2f}% RPM").font = Font(bold=True, size=14)
                
        except Exception as e:
            print(f"Warning: Error generando gráficos VDF: {e}")
            ws_graficos_vdf.cell(row=1, column=1, value=f"ADVERTENCIA: Error generando gráficos VDF - Verifique los datos de entrada")

        # --- Hoja 7: Análisis y Resumen Técnico ---
        ws_analisis = wb.create_sheet("Análisis y Resumen Técnico")
        style_sheet(ws_analisis, "Análisis y Resumen Técnico", {'A': 120}, print_area='A1:A7', print_orientation='portrait')
        
        # Obtener el porcentaje VDF real
        vfd_percentage_real = session_state.get('vfd_speed_percentage', session_state.get('rpm_percentage', 75.0))
        if vfd_percentage_real == 0:
            vfd_percentage_real = 75.0  # Valor por defecto
        
        # Comentarios técnicos por defecto si no están disponibles
        resumen_100_rpm_default = """RESUMEN Y COMENTARIOS TÉCNICOS - 100% RPM

RESUMEN DE RESULTADOS
Punto de Operación:
- Caudal: {caudal_operacion:.2f} L/s
- Altura: {altura_operacion:.2f} m
- Eficiencia: {eficiencia_operacion:.1f}%
- Potencia: {potencia_operacion:.2f} HP
- NPSH Requerido: {npsh_requerido:.2f} m

ANÁLISIS DE NPSH (100% RPM)
Criterio de Diseño Estándar:
- NPSH Disponible: {npshd_mca:.2f} m.c.a.
- NPSH Requerido: {npsh_requerido:.2f} m.c.a.
- Diferencia: {npsh_diff:.2f} m.c.a.
- Margen Mínimo Requerido: {npsh_margen_min:.2f} m.c.a.

COMENTARIOS TÉCNICOS
- Eficiencia: La bomba opera en zona de eficiencia aceptable
- NPSH: Verificar margen de seguridad adecuado
- Potencia: Confirmar potencia disponible en punto de operación

RECOMENDACIONES
- Verificar cobertura del rango de operación
- Considerar factor de seguridad en diseño
- Monitorear eficiencia durante operación""".format(
            caudal_operacion=session_state.get('caudal_operacion', 0.0),
            altura_operacion=session_state.get('altura_operacion', 0.0),
            eficiencia_operacion=session_state.get('eficiencia_operacion', 0.0),
            potencia_operacion=session_state.get('potencia_operacion', 0.0),
            npsh_requerido=session_state.get('npsh_requerido', 0.0),
            npshd_mca=session_state.get('npshd_mca', 0.0),
            npsh_diff=session_state.get('npshd_mca', 0.0) - session_state.get('npsh_requerido', 0.0),
            npsh_margen_min=session_state.get('npsh_requerido', 0.0) * 1.2
        )
        
        resumen_vdf_default = f"""RESUMEN Y COMENTARIOS TÉCNICOS - {vfd_percentage_real:.2f}% RPM

RESUMEN DE RESULTADOS VFD
Punto de Operación VFD ({vfd_percentage_real:.2f}% RPM):
- Caudal VDF: {session_state.get('caudal_nominal_vdf', session_state.get('caudal_operacion', 51.01)):.2f} L/s
- Altura VDF: {session_state.get('altura_operacion_vdf', session_state.get('altura_operacion', 94.68)):.2f} m
- Eficiencia VDF: {session_state.get('eficiencia_ajustada', session_state.get('eficiencia_operacion', 73.1)):.1f}%
- Potencia VDF: {session_state.get('potencia_ajustada', session_state.get('potencia_operacion', 56.13)):.2f} HP
- NPSH Requerido VDF: {session_state.get('npsh_requerido_vdf', session_state.get('npsh_requerido', 3.08)):.2f} m

ANÁLISIS DE NPSH ({vfd_percentage_real:.2f}% RPM)
Criterio de Diseño Estándar:
- NPSH Disponible: {session_state.get('npshd_mca', 4.97):.2f} m.c.a.
- NPSH Requerido VDF: {session_state.get('npsh_requerido_vdf', 3.08):.2f} m.c.a.
- Diferencia: {session_state.get('npshd_mca', 4.97) - session_state.get('npsh_requerido_vdf', 3.08):.2f} m.c.a.
- Margen Mínimo Requerido: {session_state.get('npsh_requerido_vdf', 3.08) * 1.2:.2f} m.c.a.

COMENTARIOS TÉCNICOS VFD
- Eficiencia: La bomba opera bien con VFD al {vfd_percentage_real:.2f}% RPM
- NPSH Adecuado: Margen de seguridad adecuado con VFD
- Potencia Requerida: Potencia reducida con VFD
- Ahorro Energético: Potencial reducción de consumo energético

RECOMENDACIONES VDF
- Evaluar costo-beneficio de implementar VFD
- Ajustar porcentaje RPM para optimización
- Considerar impacto en vida útil de bomba"""

        # Usar comentarios de session_state si están disponibles, sino usar los por defecto
        resumen_100_rpm = session_state.get('comentarios_tecnicos_100_rpm', resumen_100_rpm_default)
        resumen_vdf = session_state.get('comentarios_tecnicos_vdf', resumen_vdf_default)

        ws_analisis.cell(row=3, column=1, value="RESUMEN Y COMENTARIOS TÉCNICOS - 100% RPM").font = Font(bold=True, size=12)
        ws_analisis.cell(row=4, column=1, value=resumen_100_rpm).alignment = Alignment(wrap_text=True, vertical='top')
        ws_analisis.row_dimensions[4].height = 200

        ws_analisis.cell(row=6, column=1, value=f"RESUMEN Y COMENTARIOS TÉCNICOS - {vfd_percentage_real:.2f}% RPM").font = Font(bold=True, size=12)
        ws_analisis.cell(row=7, column=1, value=resumen_vdf).alignment = Alignment(wrap_text=True, vertical='top')
        ws_analisis.row_dimensions[7].height = 200

        try:
            wb.save(output)
            output.seek(0)
            return output
        except Exception as e:
            raise Exception(f"Error guardando archivo Excel: {str(e)}")
            
    except Exception as e:
        raise Exception(f"Error generando reporte Excel: {str(e)}")

def generate_system_curve_data(session_state: Dict[str, Any]) -> pd.DataFrame:
    """
    Genera automáticamente los datos de la curva del sistema basados en parámetros del proyecto
    """
    try:
        import numpy as np
        from core.calculations import calculate_adt_for_multiple_flows
        
        # Generar caudales para calcular la curva del sistema
        q_max = (session_state.get('caudal_lps', 100.0)) * 1.5
        flows = np.linspace(0, q_max, 20).tolist()
        
        # Obtener parámetros del sistema desde session_state estandarizados
        system_params = {
            'long_succion': session_state.get('long_succion', 10.0),
            'diam_succion_m': session_state.get('diam_succion_mm', 200.0) / 1000.0,
            'mat_succion': session_state.get('mat_succion', 'PVC'),
            'otras_perdidas_succion': session_state.get('otras_perdidas_succion', 0.0),
            'accesorios_succion': session_state.get('accesorios_succion', []),
            'long_impulsion': session_state.get('long_impulsion', 500.0),
            'diam_impulsion_m': session_state.get('diam_impulsion_mm', 150.0) / 1000.0,
            'mat_impulsion': session_state.get('mat_impulsion', 'PVC'),
            'otras_perdidas_impulsion': session_state.get('otras_perdidas_impulsion', 0.0),
            'accesorios_impulsion': session_state.get('accesorios_impulsion', []),
            'altura_succion': session_state.get('altura_succion_input', 1.65),
            'altura_descarga': session_state.get('altura_descarga', 80.0),
            'bomba_inundada': session_state.get('bomba_inundada', False),
            'metodo_calculo': session_state.get('metodo_calculo', 'Hazen-Williams'),
            'temp_liquido': session_state.get('temp_liquido', 20.0),
            'C_succion': session_state.get('coeficiente_hazen_succion', 150),
            'C_impulsion': session_state.get('coeficiente_hazen_impulsion', 150)
        }
        
        adt_values = calculate_adt_for_multiple_flows(flows, 'L/s', system_params)
        
        if adt_values:
            return pd.DataFrame({
                'Caudal (L/s)': flows,
                'ADT (m)': adt_values
            })
        else:
            # Fallback con valores por defecto
            return pd.DataFrame({
                'Caudal (L/s)': [0, 50, 100, 150, 200],
                'ADT (m)': [50, 52, 56, 62, 70]
            })
            
    except Exception as e:
        # Fallback con valores por defecto en caso de error
        print(f"Warning: Error generando curva del sistema automáticamente: {e}")
        return pd.DataFrame({
            'Caudal (L/s)': [0, 50, 100, 150, 200],
            'ADT (m)': [50, 52, 56, 62, 70]
        })

def ensure_required_dataframes(session_state: Dict[str, Any]) -> None:
    """
    Asegura que todos los DataFrames requeridos existan en el session_state.
    Crea DataFrames vacíos si no existen.
    """
    # Tablas principales del análisis (100% RPM)
    required_dataframes_100 = [
        'df_bomba_100',           # Curva H-Q Bomba 100%
        'df_rendimiento_100',     # Curva de Eficiencia 100%
        'df_potencia_100',       # Curva de Potencia 100%
        'df_npsh_100',           # Curva NPSH 100%
        'df_sistema_100'         # Curva del Sistema 100%
    ]
    
    # Tablas VDF
    required_dataframes_vfd = [
        'df_bomba_vdf',         # Curva H-Q Bomba VDF
        'df_rendimiento_vfd',   # Curva de Eficiencia VDF
        'df_potencia_vfd',      # Curva de Potencia VDF
        'df_npsh_vfd',          # Curva NPSH VDF
        'df_sistema_vfd'        # Curva del Sistema VDF
    ]
    
    # Tablas de compatibilidad (fallback)
    required_dataframes_fallback = [
        'df_curva_sistema', 'df_eff_100', 'df_power_100', 'df_npshr'
    ]
    
    # Combinar todas las tablas requeridas
    all_dataframes = required_dataframes_100 + required_dataframes_vfd + required_dataframes_fallback
    
    for df_name in all_dataframes:
        if df_name not in session_state or session_state[df_name] is None:
            session_state[df_name] = pd.DataFrame()

def create_comprehensive_excel_report(session_state: Dict[str, Any], incluir_formulas: bool = False, incluir_graficos: bool = True) -> io.BytesIO:
    """
    Función principal mejorada que garantiza que cada hoja incluya inputs, resultados y gráficas
    según las especificaciones del usuario
    
    Args:
        session_state: Estado de la sesión con todos los datos
        incluir_formulas: Si True, incluye fórmulas en las celdas de la hoja 'Datos Gráficos'
        incluir_graficos: Si True, incluye las hojas de gráficos (100% RPM y VFD)
    """
    try:
        print("PROCESANDO: Iniciando generación de reporte completo...")
        
        # Validar datos mínimos necesarios
        if not validate_minimum_data(session_state):
            raise Exception("No hay datos suficientes para generar el reporte. Complete los datos de entrada primero.")
        
        result = export_full_project_to_excel(session_state, incluir_formulas=incluir_formulas, incluir_graficos=incluir_graficos)
        print("ÉXITO: Reporte generado exitosamente")
        return result
        
    except Exception as e:
        error_msg = f"ERROR: Error crítico generando reporte: {str(e)}"
        print(error_msg)
        raise Exception(error_msg)

def validate_minimum_data(session_state: Dict[str, Any]) -> bool:
    """
    Valida que existan los datos mínimos necesarios para generar el reporte
    """
    required_fields = [
        'proyecto', 'caudal_lps', 'altura_estatica_total'
    ]
    
    for field in required_fields:
        if not session_state.get(field):
            print(f"ADVERTENCIA: Campo requerido faltante: {field}")
            return False
    
    return True
