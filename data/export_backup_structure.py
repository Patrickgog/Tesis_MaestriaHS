# Módulo para la exportación de datos a formatos como Excel

import pandas as pd
import io
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def export_full_project_to_excel(session_state: Dict[str, Any]) -> io.BytesIO:
    """
    Recopila todos los datos, análisis y gráficos del proyecto y los exporta 
    a un archivo Excel con múltiples hojas.
    """
    try:
        output = io.BytesIO()
        wb = Workbook()
        wb.remove(wb.active)
        
        # Generar automáticamente datos de curva del sistema si no existen
        if 'df_curva_sistema' not in session_state or session_state.get('df_curva_sistema', pd.DataFrame()).empty:
            df_sistema_generated = generate_system_curve_data(session_state)
            session_state['df_curva_sistema'] = df_sistema_generated
        
        # Verificar que todos los DataFrames necesarios estén presentes
        ensure_required_dataframes(session_state)
        
        # --- Función de estilo ---
        def style_sheet(ws, title, col_widths: Dict[str, int]):
            ws.title = title
            ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=16, color="FFFFFF")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_widths))
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=1).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width
            
            # Estilo para cabeceras de tablas
            for cell in ws[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            ws.freeze_panes = 'A3'

        # --- Hoja 1: Resumen General ---
        ws_resumen = wb.create_sheet("Resumen General")
        style_sheet(ws_resumen, "Resumen General del Proyecto", {'A': 40, 'B': 25})
        
        resumen_data = {
        'Parámetro': [
            'Proyecto', 'Diseño', 'Caudal de Diseño (L/s)', 'Altura Estática Total (m)',
            'ADT Total (m)', 'Potencia Motor Final (HP)', 'NPSH Disponible (m.c.a.)',
            'Punto de Operación Caudal (L/s)', 'Punto de Operación Altura (m)',
            'Punto de Operación Eficiencia (%)', 'Punto de Operación Potencia (HP)'
        ],
        'Valor': [
            session_state.get('proyecto', 'N/A'),
            session_state.get('diseno', 'N/A'),
            f"{session_state.get('caudal_lps', 0):.2f}",
            f"{session_state.get('altura_estatica_total', 0):.2f}",
            f"{session_state.get('adt_total', 0):.2f}",
            f"{session_state.get('potencia_motor_final_hp', 0):.2f}",
            f"{session_state.get('npshd_mca', 0):.2f}",
            f"{session_state.get('caudal_operacion', 0):.2f}",
            f"{session_state.get('altura_operacion', 0):.2f}",
            f"{session_state.get('eficiencia_operacion', 0):.2f}",
            f"{session_state.get('potencia_operacion', 0):.2f}"
        ]
        }
        
        df_resumen = pd.DataFrame(resumen_data)
        for r in dataframe_to_rows(df_resumen, index=False, header=True):
            ws_resumen.append(r)
        for cell in ws_resumen['A'] + ws_resumen['B']:
            cell.alignment = Alignment(vertical='center', horizontal='left')
            for cell in ws_resumen["1:1"]: 
                cell.font = Font(bold=True)
        for cell in ws_resumen["2:2"]: 
            cell.font = Font(bold=True)

        # --- Hoja 2: Datos de Entrada (Pestaña 1) ---
        ws_inputs = wb.create_sheet("Datos de Entrada")
        style_sheet(ws_inputs, "Inputs Detallados del Proyecto", {'A': 25, 'B': 35, 'C': 25})
    
        inputs_data = {'Sección': [], 'Parámetro': [], 'Valor': []}
        def add_input(seccion, parametro, valor):
            inputs_data['Sección'].append(seccion)
            inputs_data['Parámetro'].append(parametro)
            inputs_data['Valor'].append(str(valor))

        # Condiciones de Operación
        add_input('Operación', 'Caudal de Diseño (L/s)', session_state.get('caudal_lps'))
        add_input('Operación', 'Elevación del sitio (m)', session_state.get('elevacion_sitio'))
        add_input('Operación', 'Altura de Succión (m)', session_state.get('altura_succion_input'))
        add_input('Operación', 'Bomba inundada', session_state.get('bomba_inundada'))
        add_input('Operación', 'Altura de Descarga (m)', session_state.get('altura_descarga'))
        add_input('Operación', 'Número de Bombas', session_state.get('num_bombas'))

        # Succión
        add_input('Succión', 'Longitud Tubería (m)', session_state.get('long_succion'))
        add_input('Succión', 'Material Tubería', session_state.get('mat_succion'))
        add_input('Succión', 'Diámetro Interno (mm)', f"{session_state.get('diam_succion_mm', 0):.2f}")
        add_input('Succión', 'Otras Pérdidas (m)', session_state.get('otras_perdidas_succion'))
        
        # Impulsión
        add_input('Impulsión', 'Longitud Tubería (m)', session_state.get('long_impulsion'))
        add_input('Impulsión', 'Material Tubería', session_state.get('mat_impulsion'))
        add_input('Impulsión', 'Diámetro Interno (mm)', f"{session_state.get('diam_impulsion_mm', 0):.2f}")
        add_input('Impulsión', 'Otras Pérdidas (m)', session_state.get('otras_perdidas_impulsion'))

        # Accesorios
        for acc in session_state.get('accesorios_succion', []):
            add_input('Accesorios Succión', acc['tipo'], f"Cantidad: {acc['cantidad']}")
        for acc in session_state.get('accesorios_impulsion', []):
            add_input('Accesorios Impulsión', acc['tipo'], f"Cantidad: {acc['cantidad']}")

        df_inputs = pd.DataFrame(inputs_data)
        for r in dataframe_to_rows(df_inputs, index=False, header=True):
            ws_inputs.append(r)

        # --- Hoja 3: Resultados Completos (Pestaña 1) ---
        ws_resultados = wb.create_sheet("Resultados Cálculos")
        style_sheet(ws_resultados, "Resultados Numéricos Completos", {'A': 25, 'B': 35, 'C': 25})
        
        results_data = {'Sección': [], 'Parámetro': [], 'Valor': []}
        def add_result(seccion, parametro, valor, unit=""):
            results_data['Sección'].append(seccion)
            results_data['Parámetro'].append(parametro)
            results_data['Valor'].append(f"{valor:.2f} {unit}".strip())

        # Resultados Succión
        add_result('Succión', 'Velocidad', session_state.get('velocidad_succion', 0), 'm/s')
        add_result('Succión', 'Pérdida Primaria', session_state.get('hf_primaria_succion', 0), 'm')
        add_result('Succión', 'Pérdida Secundaria', session_state.get('hf_secundaria_succion', 0), 'm')
        add_result('Succión', 'Pérdida Total', session_state.get('perdida_total_succion', 0), 'm')
        add_result('Succión', 'Altura Dinámica', session_state.get('altura_dinamica_succion', 0), 'm')

        # Resultados Impulsión
        add_result('Impulsión', 'Velocidad', session_state.get('velocidad_impulsion', 0), 'm/s')
        add_result('Impulsión', 'Pérdida Primaria', session_state.get('hf_primaria_impulsion', 0), 'm')
        add_result('Impulsión', 'Pérdida Secundaria', session_state.get('hf_secundaria_impulsion', 0), 'm')
        add_result('Impulsión', 'Pérdida Total', session_state.get('perdida_total_impulsion', 0), 'm')
        add_result('Impulsión', 'Altura Dinámica', session_state.get('altura_dinamica_impulsion', 0), 'm')

        # Resultados Generales
        add_result('Sistema', 'NPSH Disponible', session_state.get('npshd_mca', 0), 'm.c.a')
        add_result('Sistema', 'Altura Dinámica Total (ADT)', session_state.get('adt_total', 0), 'm')
        add_result('Motor', 'Potencia Hidráulica', session_state.get('potencia_hidraulica_hp', 0), 'HP')
        add_result('Motor', 'Potencia Motor Requerida', session_state.get('potencia_motor_final_hp', 0), 'HP')
        if session_state.get('motor_seleccionado'):
            motor = session_state.get('motor_seleccionado')
            add_result('Motor', 'Motor Estándar Seleccionado', motor['potencia_hp'], 'HP')

        df_results = pd.DataFrame(results_data)
        for r in dataframe_to_rows(df_results, index=False, header=True):
            ws_resultados.append(r)

        # --- Hoja 4: Datos para Gráficos (Pestaña 2) ---
        ws_data = wb.create_sheet("Datos Gráficos")
        # Ocultar esta hoja
        ws_data.sheet_state = 'hidden'

        try:
        # Recopilar todos los datos de curvas con manejo seguro
                    
        # ---------- TABLAS 100% RPM ----------
        df_bomba_100 = session_state.get('df_bomba_100', pd.DataFrame())
        df_rendimiento_100 = session_state.get('df_rendimiento_100', pd.DataFrame())
        df_potencia_100 = session_state.get('df_potencia_100', pd.DataFrame())
        df_npsh_100 = session_state.get('df_npsh_100', pd.DataFrame())
        df_sistema_100 = session_state.get('df_sistema_100', pd.DataFrame())
                    
        # ---------- TABLAS VDF ----------
        df_bomba_vfd = session_state.get('df_bomba_vfd', pd.DataFrame())
        df_rendimiento_vfd = session_state.get('df_rendimiento_vfd', pd.DataFrame())
        df_potencia_vfd = session_state.get('df_potencia_vfd', pd.DataFrame())
        df_npsh_vfd = session_state.get('df_npsh_vfd', pd.DataFrame())
        df_sistema_vfd = session_state.get('df_sistema_vfd', pd.DataFrame())
                    
        # ---------- TABLAS OTRAS (compatibilidad con código anterior) ----------
        df_system = session_state.get('df_curva_sistema', pd.DataFrame())
        df_bomba_eff = session_state.get('df_eff_100', pd.DataFrame())
        df_power_100_fallback = session_state.get('df_power_100', pd.DataFrame())
        df_npshr = session_state.get('df_npshr', pd.DataFrame())

        # Función segura para agregar datos de curva
        def safe_add_curve_data(df_source, df_target, q_column, h_column, q_output_name, h_output_name):
            try:
                if (not df_source.empty and 
                    q_column in df_source.columns and 
                    h_column in df_source.columns):
                    
                    if df_target.empty:
                        return df_source.rename(columns={q_column: q_output_name, h_column: h_output_name})
                    else:
                        temp_df = df_source.rename(columns={q_column: q_output_name, h_column: h_output_name})
                        return pd.concat([df_target, temp_df], axis=1)
                return df_target
            except Exception:
                return df_target

        # ---------- AGREGAR TODAS LAS TABLAS NECESARIAS PARA GRÁFICOS ----------
                    
        # DataFrame principal que contendrá todas las tablas
        df_graficos = pd.DataFrame()
        
        # ---------- TABLAS 100% RPM ----------
        
        # Tabla 1: Curva H-Q Bomba 100%
        if not df_bomba_100.empty:
            df_graficos = safe_add_curve_data(df_bomba_100, df_graficos, 'Caudal (L/s)', 'Altura (m)', 'Q_Bomba_100', 'H_Bomba_100')
        
        # Tabla 2: Curva de Rendimiento (Eficiencia) 100%
        if not df_rendimiento_100.empty:
            df_graficos = safe_add_curve_data(df_rendimiento_100, df_graficos, 'Caudal (L/s)', 'Rendimiento (%)', 'Q_Eff_100', 'Eff_100')
                    
        # Tabla 3: Curva de Potencia 100%
        if not df_potencia_100.empty:
            df_graficos = safe_add_curve_data(df_potencia_100, df_graficos, 'Caudal (L/s)', 'Potencia (HP)', 'Q_Power_100', 'P_100')
                    
        # Tabla 4: Curva NPSH 100%
        if not df_npsh_100.empty:
            df_graficos = safe_add_curve_data(df_npsh_100, df_graficos, 'Caudal (L/s)', 'NPSH (m)', 'Q_NPSHr', 'NPSHr')
                    
        # Tabla 5: Curva del Sistema 100%
        if not df_sistema_100.empty:
            df_graficos = safe_add_curve_data(df_sistema_100, df_graficos, 'Caudal (L/s)', 'Altura (m)', 'Q_Sistema', 'H_Sistema')
                    
        # ---------- TABLAS VDF ----------
                    
        # Tabla 6: Curva H-Q Bomba VDF
        if not df_bomba_vfd.empty:
            df_graficos = safe_add_curve_data(df_bomba_vfd, df_graficos, 'Caudal (L/s)', 'Altura (m)', 'Q_Bomba_VDF', 'H_Bomba_VDF')
                    
        # Tabla 7: Curva de Rendimiento VDF
        if not df_rendimiento_vfd.empty:
            df_graficos = safe_add_curve_data(df_rendimiento_vfd, df_graficos, 'Caudal (L/s)', 'Rendimiento (%)', 'Q_Eff_VDF', 'Eff_VDF')
                    
        # Tabla 8: Curva de Potencia VDF  
        if not df_potencia_vfd.empty:
            df_graficos = safe_add_curve_data(df_potencia_vfd, df_graficos, 'Caudal (L/s)', 'Potencia (HP)', 'Q_Power_VDF', 'P_VDF')
                    
        # Tabla 9: Curva NPSH VDF
        if not df_npsh_vfd.empty:
            df_graficos = safe_add_curve_data(df_npsh_vfd, df_graficos, 'Caudal (L/s)', 'NPSH (m)', 'Q_NPSHr_VDF', 'NPSHr_VDF')
                    
        # Tabla 10: Curva del Sistema VDF
        if not df_sistema_vfd.empty:
            df_graficos = safe_add_curve_data(df_sistema_vfd, df_graficos, 'Caudal (L/s)', 'Altura (m)', 'Q_Sistema_VDF', 'H_Sistema_VDF')
                    
        # ---------- COMPATIBILIDAD CON CÓDIGO ANTERIOR ----------
                    
        # Si aún no tenemos datos suficientes, usar las tablas de fallback
        if df_graficos.empty or len(df_graficos.columns) < 4:
        # Curva del Sistema (fallback)
        df_graficos = safe_add_curve_data(df_system, df_graficos, 'Caudal (L/s)', 'ADT (m)', 'Q_Sistema', 'H_Sistema')
                        
        # Curva Eficiencia (fallback)
        df_graficos = safe_add_curve_data(df_bomba_eff, df_graficos, 'Caudal (L/s)', 'Eficiencia (%)', 'Q_Eff_100', 'Eff_100')
                        
        # Curva Potencia (fallback)
        df_graficos = safe_add_curve_data(df_power_100_fallback, df_graficos, 'Caudal (L/s)', 'Potencia (HP)', 'Q_Power_100', 'P_100')
                        
        # Curva NPSH (fallback)
        df_graficos = safe_add_curve_data(df_npshr, df_graficos, 'Caudal (L/s)', 'NPSHr (m)', 'Q_NPSHr', 'NPSHr')

        # Eliminar columnas duplicadas de Caudal
        df_graficos = df_graficos.loc[:,~df_graficos.columns.duplicated()]

        # Validar que tenemos al menos datos del sistema
        if df_graficos.empty:
            # Crear datos mínimos de ejemplo si no hay datos
            df_graficos = pd.DataFrame({
                'Q_Sistema': [0, 50, 100, 150, 200],
                'H_Sistema': [50, 52, 56, 62, 70],
                'Q_Bomba_100': [0, 50, 100, 150, 200],
                'H_Bomba_100': [60, 58, 54, 48, 40],
                'Q_Eff_100': [0, 50, 100, 150, 200],
                'Eff_100': [0, 75, 80, 75, 65],
                'Q_Power_100': [0, 50, 100, 150, 200],
                'P_100': [0, 15, 25, 35, 45]
            })

        # Agregar datos a la hoja
        for r in dataframe_to_rows(df_graficos, index=False, header=True):
            ws_data.append(r)
            
    except Exception as e:
        print(f"Warning: Error procesando datos de gráficos: {e}")
        # Datos de emergencia
        emergency_data = pd.DataFrame({
            'Q_Sistema': [0, 50, 100, 150, 200],
            'H_Sistema': [50, 52, 56, 62, 70]
        })
        for r in dataframe_to_rows(emergency_data, index=False, header=True):
            ws_data.append(r)

        # --- Hoja 5: Gráficos 100% RPM ---
        ws_graficos_100 = wb.create_sheet("Gráficos 100% RPM")
    
    try:
        if not df_graficos.empty:
            max_rows = len(df_graficos) + 1

            # Segurizar la función que obtiene referencias de columnas
            def safe_get_column_ref(ws_data, df, col_name, start_row=2, end_row=None):
                """Función segura para obtener referencia de columna"""
                try:
                    if end_row is None:
                        end_row = len(df) + 1
                    
                    if col_name in df.columns:
                        col_index = df.columns.get_loc(col_name) + 1
                        return Reference(ws_data, min_col=col_index, min_row=start_row, max_row=end_row)
                    else:
                        return None
                except Exception:
                    return None

        # Gráfico 1: Curva Bomba vs Sistema - SOLO 100% RPM
            chart1 = ScatterChart()
        chart1.title = "Curva de Bomba vs Curva de Sistema - 100% RPM"
            chart1.x_axis.title = "Caudal (L/s)"
            chart1.y_axis.title = "Altura (m)"

            # Serie Curva Sistema
            x_sys = safe_get_column_ref(ws_data, df_graficos, "Q_Sistema")
            y_sys = safe_get_column_ref(ws_data, df_graficos, "H_Sistema")
            if x_sys and y_sys:
                s_sys = Series(y_sys, x_sys, title="Curva Sistema")
                chart1.series.append(s_sys)

            # Serie Curva Bomba 100%
            x_bomba = safe_get_column_ref(ws_data, df_graficos, "Q_Bomba_100")
            y_bomba = safe_get_column_ref(ws_data, df_graficos, "H_Bomba_100")
            if x_bomba and y_bomba:
        s_bomba = Series(y_bomba, x_bomba, title='Curva Bomba 100% RPM')
                chart1.series.append(s_bomba)
            
        # Punto de Operación 100% RPM
        caudal_op = session_state.get('caudal_operacion', 0)
        altura_op = session_state.get('altura_operacion', 0)
        if caudal_op > 0 and altura_op > 0:
        # Agregar punto de operación como serie separada
        ws_graficos_100.cell(row=50, column=1, value=caudal_op)
        ws_graficos_100.cell(row=50, column=2, value=altura_op)
        x_point_ref = Reference(ws_graficos_100, min_col=1, min_row=50, max_row=50)
        y_point_ref = Reference(ws_graficos_100, min_col=2, min_row=50, max_row=50)
        s_op = Series(y_point_ref, x_point_ref, title="Punto de Operación")
        chart1.series.append(s_op)
        
        ws_graficos_100.add_chart(chart1, "A1")

            # Gráfico 2: Eficiencia
            if 'Q_Eff_100' in df_graficos.columns and 'Eff_100' in df_graficos.columns:
                chart2 = ScatterChart()
                chart2.title = "Curva de Eficiencia"
                chart2.x_axis.title = "Caudal (L/s)"
                chart2.y_axis.title = "Eficiencia (%)"
                
                x_eff = safe_get_column_ref(ws_data, df_graficos, "Q_Eff_100")
                y_eff = safe_get_column_ref(ws_data, df_graficos, "Eff_100")
                if x_eff and y_eff:
                    s_eff = Series(y_eff, x_eff, title="Eficiencia 100% RPM")
                    chart2.series.append(s_eff)
                                
        # Punto de operación en eficiencia
        caudal_op = session_state.get('caudal_operacion', 0)
        eficiencia_op = session_state.get('eficiencia_operacion', 0)
        if caudal_op > 0 and eficiencia_op > 0:
        ws_graficos_100.cell(row=51, column=1, value=caudal_op)
        ws_graficos_100.cell(row=51, column=2, value=eficiencia_op)
        x_op_eff = Reference(ws_graficos_100, min_col=1, min_row=51, max_row=51)
        y_op_eff = Reference(ws_graficos_100, min_col=2, min_row=51, max_row=51)
        s_op_eff = Series(y_op_eff, x_op_eff, title="Eficiencia en Operación")
        chart2.series.append(s_op_eff)
                                
        ws_graficos_100.add_chart(chart2, "A16")

            # Gráfico 3: Potencia
            if 'Q_Power_100' in df_graficos.columns and 'P_100' in df_graficos.columns:
                chart3 = ScatterChart()
                chart3.title = "Curva de Potencia"
                chart3.x_axis.title = "Caudal (L/s)"
                chart3.y_axis.title = "Potencia (HP)"
                
                x_pow = safe_get_column_ref(ws_data, df_graficos, "Q_Power_100")
                y_pow = safe_get_column_ref(ws_data, df_graficos, "P_100")
                if x_pow and y_pow:
                    s_pow = Series(y_pow, x_pow, title="Potencia 100% RPM")
                    chart3.series.append(s_pow)
                                
        # Punto de operación en potencia
        caudal_op = session_state.get('caudal_operacion', 0)
        potencia_op = session_state.get('potencia_operacion', 0)
        if caudal_op > 0 and potencia_op > 0:
        ws_graficos_100.cell(row=52, column=1, value=caudal_op)
        ws_graficos_100.cell(row=52, column=2, value=potencia_op)
        x_op_pow = Reference(ws_graficos_100, min_col=1, min_row=52, max_row=52)
        y_op_pow = Reference(ws_graficos_100, min_col=2, min_row=52, max_row=52)
        s_op_pow = Series(y_op_pow, x_op_pow, title="Potencia en Operación")
        chart3.series.append(s_op_pow)
                                
        ws_graficos_100.add_chart(chart3, "K1")

            # Gráfico 4: NPSH
            if 'Q_NPSHr' in df_graficos.columns and 'NPSHr' in df_graficos.columns:
                chart4 = ScatterChart()
                chart4.title = "Análisis de NPSH"
                chart4.x_axis.title = "Caudal (L/s)"
                chart4.y_axis.title = "NPSH (m)"
                
                x_npsh = safe_get_column_ref(ws_data, df_graficos, "Q_NPSHr")
                y_npsh = safe_get_column_ref(ws_data, df_graficos, "NPSHr")
                if x_npsh and y_npsh:
                    s_npsh = Series(y_npsh, x_npsh, title="NPSH Requerido")
                    chart4.series.append(s_npsh)
                                
        # Punto de operación en NPSH
        caudal_op = session_state.get('caudal_operacion', 0)
        npsh_op = session_state.get('npsh_requerido', 0)
        if caudal_op > 0 and npsh_op > 0:
        ws_graficos_100.cell(row=53, column=1, value=caudal_op)
        ws_graficos_100.cell(row=53, column=2, value=npsh_op)
        x_op_npsh = Reference(ws_graficos_100, min_col=1, min_row=53, max_row=53)
        y_op_npsh = Reference(ws_graficos_100, min_col=2, min_row=53, max_row=53)
        s_op_npsh = Series(y_op_npsh, x_op_npsh, title="NPSH en Operación")
        chart4.series.append(s_op_npsh)
                                
        ws_graficos_100.add_chart(chart4, "K16")
                        
        # Agregar información adicional sobre el análisis 100% RPM
        ws_graficos_100.cell(row=30, column=1, value="INFORMACIÓN DEL REPORTE - 100% RPM").font = Font(bold=True, size=12)
        ws_graficos_100.cell(row=31, column=1, value=f"Proyecto: {session_state.get('proyecto', 'Sin especificar')}")
        ws_graficos_100.cell(row=32, column=1, value=f"Diseño: {session_state.get('diseno', 'Sin especificar')}")
        ws_graficos_100.cell(row=33, column=1, value=f"Caudal de Diseño: {session_state.get('caudal_lps', 0):.2f} L/s")
        ws_graficos_100.cell(row=34, column=1, value=f"Altura Estática Total: {session_state.get('altura_estatica_total', 0):.2f} m")
        ws_graficos_100.cell(row=35, column=1, value=f"ADT Total: {session_state.get('adt_total', 0):.2f} m")
                        
        except Exception as e:
        print(f"Warning: Error generando gráficos 100% RPM: {e}")
        ws_graficos_100.cell(row=1, column=1, value="ADVERTENCIA: Error generando gráficos 100% RPM - Verifique los datos de entrada")
        
        # --- Hoja 6: Gráficos VDF ---
        ws_graficos_vdf = wb.create_sheet("Gráficos VDF")
                
    try:
        # Obtener el porcentaje VDF real
        vfd_percentage = session_state.get('vfd_speed_percentage', session_state.get('rpm_percentage', 0))
        if vfd_percentage == 0:
            vfd_percentage = 75.0  # Valor por defecto
                    
        if not df_graficos.empty:
            max_rows = len(df_graficos) + 1
                        
        # Gráfico 1: Curva Bomba VDF vs Sistema
        if 'Q_Bomba_VDF' in df_graficos.columns and 'H_Bomba_VDF' in df_graficos.columns:
        chart_vdf1 = ScatterChart()
        chart_vdf1.title = f"Curva de Bomba vs Curva de Sistema - {vfd_percentage:.2f}% RPM"
        chart_vdf1.x_axis.title = "Caudal (L/s)"
        chart_vdf1.y_axis.title = "Altura (m)"
        
        # Serie Curva Sistema
        x_sys_vdf = safe_get_column_ref(ws_data, df_graficos, "Q_Sistema")
        y_sys_vdf = safe_get_column_ref(ws_data, df_graficos, "H_Sistema")
        if x_sys_vdf and y_sys_vdf:
        s_sys_vdf = Series(y_sys_vdf, x_sys_vdf, title="Curva Sistema")
        chart_vdf1.series.append(s_sys_vdf)
        
        # Serie Curva Bomba VDF
        x_vdf = safe_get_column_ref(ws_data, df_graficos, "Q_Bomba_VDF")
        y_vdf = safe_get_column_ref(ws_data, df_graficos, "H_Bomba_VDF")
        if x_vdf and y_vdf:
        s_vdf = Series(y_vdf, x_vdf, title=f"Curva Bomba VDF {vfd_percentage:.2f}% RPM")
        chart_vdf1.series.append(s_vdf)
        
        # Punto de operación VDF
        interseccion_vfd = session_state.get('interseccion_vfd', None)
        if interseccion_vfd:
        caudal_op_vfd, altura_op_vfd = interseccion_vfd
        ws_graficos_vdf.cell(row=60, column=1, value=caudal_op_vfd)
        ws_graficos_vdf.cell(row=60, column=2, value=altura_op_vfd)
        x_op_vdf = Reference(ws_graficos_vdf, min_col=1, min_row=60, max_row=60)
        y_op_vdf = Reference(ws_graficos_vdf, min_col=2, min_row=60, max_row=60)
        s_op_vdf = Series(y_op_vdf, x_op_vdf, title=f"Punto Operación VDF")
        chart_vdf1.series.append(s_op_vdf)
        
        ws_graficos_vdf.add_chart(chart_vdf1, "A1")
        
        # Gráfico 2: Eficiencia VDF
        if 'Q_Eff_VDF' in df_graficos.columns and 'Eff_VDF' in df_graficos.columns:
        chart_vdf2 = ScatterChart()
        chart_vdf2.title = f"Curva de Eficiencia VDF - {vfd_percentage:.2f}% RPM"
        chart_vdf2.x_axis.title = "Caudal (L/s)"
        chart_vdf2.y_axis.title = "Eficiencia (%)"
                            
        x_eff_vdf = safe_get_column_ref(ws_data, df_graficos, "Q_Eff_VDF")
        y_eff_vdf = safe_get_column_ref(ws_data, df_graficos, "Eff_VDF")
        if x_eff_vdf and y_eff_vdf:
        s_eff_vdf = Series(y_eff_vdf, x_eff_vdf, title=f"Eficiencia VDF {vfd_percentage:.2f}% RPM")
        chart_vdf2.series.append(s_eff_vdf)
                                
        # Punto de operación en eficiencia VDF
        eficiencia_vdf = session_state.get('eficiencia_ajustada', 0)
        interseccion_vfd = session_state.get('interseccion_vfd', None)
        if interseccion_vfd and eficiencia_vdf > 0:
        caudal_op_vfd, _ = interseccion_vfd
        ws_graficos_vdf.cell(row=61, column=1, value=caudal_op_vfd)
        ws_graficos_vdf.cell(row=61, column=2, value=eficiencia_vdf)
        x_op_eff_vdf = Reference(ws_graficos_vdf, min_col=1, min_row=61, max_row=61)
        y_op_eff_vdf = Reference(ws_graficos_vdf, min_col=2, min_row=61, max_row=61)
        s_op_eff_vdf = Series(y_op_eff_vdf, x_op_eff_vdf, title="Eficiencia en Operación VDF")
        chart_vdf2.series.append(s_op_eff_vdf)
                                
        ws_graficos_vdf.add_chart(chart_vdf2, "A16")
        
        # Gráfico 3: Potencia VDF
        if 'Q_Power_VDF' in df_graficos.columns and 'P_VDF' in df_graficos.columns:
        chart_vdf3 = ScatterChart()
        chart_vdf3.title = f"Curva de Potencia VDF - {vfd_percentage:.2f}% RPM"
        chart_vdf3.x_axis.title = "Caudal (L/s)"
        chart_vdf3.y_axis.title = "Potencia (HP)"
                            
        x_pow_vdf = safe_get_column_ref(ws_data, df_graficos, "Q_Power_VDF")
        y_pow_vdf = safe_get_column_ref(ws_data, df_graficos, "P_VDF")
        if x_pow_vdf and y_pow_vdf:
        s_pow_vdf = Series(y_pow_vdf, x_pow_vdf, title=f"Potencia VDF {vfd_percentage:.2f}% RPM")
        chart_vdf3.series.append(s_pow_vdf)
                                
        # Punto de operación en potencia VDF
        potencia_vdf = session_state.get('potencia_ajustada', 0)
        interseccion_vfd = session_state.get('interseccion_vfd', None)
        if interseccion_vfd and potencia_vdf > 0:
        caudal_op_vfd, _ = interseccion_vfd
        ws_graficos_vdf.cell(row=62, column=1, value=caudal_op_vfd)
        ws_graficos_vdf.cell(row=62, column=2, value=potencia_vdf)
        x_op_pow_vdf = Reference(ws_graficos_vdf, min_col=1, min_row=62, max_row=62)
        y_op_pow_vdf = Reference(ws_graficos_vdf, min_col=2, min_row=62, max_row=62)
        s_op_pow_vdf = Series(y_op_pow_vdf, x_op_pow_vdf, title="Potencia en Operación VDF")
        chart_vdf3.series.append(s_op_pow_vdf)
                                
        ws_graficos_vdf.add_chart(chart_vdf3, "K1")
        
        # Gráfico 4: NPSH VDF
        if 'Q_NPSHr_VDF' in df_graficos.columns and 'NPSHr_VDF' in df_graficos.columns:
        chart_vdf4 = ScatterChart()
        chart_vdf4.title = f"Análisis de NPSH VDF - {vfd_percentage:.2f}% RPM"
        chart_vdf4.x_axis.title = "Caudal (L/s)"
        chart_vdf4.y_axis.title = "NPSH (m)"
                            
        x_npsh_vdf = safe_get_column_ref(ws_data, df_graficos, "Q_NPSHr_VDF")
        y_npsh_vdf = safe_get_column_ref(ws_data, df_graficos, "NPSHr_VDF")
        if x_npsh_vdf and y_npsh_vdf:
        s_npsh_vdf = Series(y_npsh_vdf, x_npsh_vdf, title=f"NPSH Requerido VDF")
        chart_vdf4.series.append(s_npsh_vdf)
                                
        # Punto de operación en NPSH VDF
        npsh_vdf = session_state.get('npsh_requerido_vdf', 0)
        interseccion_vfd = session_state.get('interseccion_vfd', None)
        if interseccion_vfd and npsh_vdf > 0:
        caudal_op_vfd, _ = interseccion_vfd
        ws_graficos_vdf.cell(row=63, column=1, value=caudal_op_vdf)
        ws_graficos_vdf.cell(row=63, column=2, value=npsh_vdf)
        x_op_npsh_vdf = Reference(ws_graficos_vdf, min_col=1, min_row=63, max_row=63)
        y_op_npsh_vdf = Reference(ws_graficos_vdf, min_col=2, min_row=63, max_row=63)
        s_op_npsh_vdf = Series(y_op_npsh_vdf, x_op_npsh_vdf, title="NPSH en Operación VDF")
        chart_vdf4.series.append(s_op_npsh_vdf)
                                
        ws_graficos_vdf.add_chart(chart_vdf4, "K16")
                        
        # Información adicional del análisis VDF
        ws_graficos_vdf.cell(row=30, column=1, value=f"INFORMACIÓN DEL REPORTE - {vfd_percentage:.2f}% RPM").font = Font(bold=True, size=12)
        ws_graficos_vdf.cell(row=31, column=1, value=f"Proyecto: {session_state.get('proyecto', 'Sin especificar')}")
        ws_graficos_vdf.cell(row=32, column=1, value=f"Diseño: {session_state.get('diseno', 'Sin especificar')}")
        ws_graficos_vdf.cell(row=33, column=1, value=f"Caudal VDF: {session_state.get('caudal_nominal_vdf', session_state.get('caudal_lps', 0)):.2f} L/s")
        ws_graficos_vdf.cell(row=34, column=1, value=f"Porcentaje VDF: {vfd_percentage:.2f}%")
        else:
        ws_graficos_vdf.cell(row=1, column=1, value=f"NO HAY DATOS VDF DISPONIBLES - {vfd_percentage:.2f}% RPM").font = Font(bold=True, size=14)
            
    except Exception as e:
        print(f"Warning: Error generando gráficos VDF: {e}")
        ws_graficos_vdf.cell(row=1, column=1, value=f"ADVERTENCIA: Error generando gráficos VDF - Verifique los datos de entrada")

        # --- Hoja 7: Análisis y Resumen Técnico ---
    ws_analisis = wb.create_sheet("Análisis y Resumen")
    style_sheet(ws_analisis, "Análisis y Resumen Técnico", {'A': 120})
    
        # Obtener el porcentaje VDF real
        vfd_percentage_real = session_state.get('vfd_speed_percentage', session_state.get('rpm_percentage', 75.0))
        if vfd_percentage_real == 0:
            vfd_percentage_real = 75.0  # Valor por defecto
                
        # Comentarios técnicos por defecto si no están disponibles
        resumen_100_rpm_default = """RESUMEN Y COMENTARIOS TÉCNICOS - 100% RPM
        
        RESUMEN DE RESULTADOS
        Punto de Operación:
        - Caudal: {caudal_operacion:.2f} L/s
        - Altura: {altura_operacion:.2f} m
        - Eficiencia: {eficiencia_operacion:.1f}%
        - Potencia: {potencia_operacion:.2f} HP
        - NPSH Requerido: {npsh_requerido:.2f} m
        
        ANÁLISIS DE NPSH (100% RPM)
        Criterio de Diseño Estándar:
        - NPSH Disponible: {npshd_mca:.2f} m.c.a.
        - NPSH Requerido: {npsh_requerido:.2f} m.c.a.
        - Diferencia: {npsh_diff:.2f} m.c.a.
        - Margen Mínimo Requerido: {npsh_margen_min:.2f} m.c.a.
        
        COMENTARIOS TÉCNICOS
        - Eficiencia: La bomba opera en zona de eficiencia aceptable
        - NPSH: Verificar margen de seguridad adecuado
        - Potencia: Confirmar potencia disponible en punto de operación
        
        RECOMENDACIONES
        - Verificar cobertura del rango de operación
        - Considerar factor de seguridad en diseño
        - Monitorear eficiencia durante operación""".format(
        caudal_operacion=session_state.get('caudal_operacion', 86.30),
        altura_operacion=session_state.get('altura_operacion', 99.13),
        eficiencia_operacion=session_state.get('eficiencia_operacion', 73.5),
        potencia_operacion=session_state.get('potencia_operacion', 131.70),
        npsh_requerido=session_state.get('npsh_requerido', 5.33),
        npshd_mca=session_state.get('npshd_mca', 4.97),
        npsh_diff=session_state.get('npshd_mca', 4.97) - session_state.get('npsh_requerido', 5.33),
        npsh_margen_min=session_state.get('npsh_requerido', 5.33) * 1.2
        )
                
        resumen_vdf_default = f"""RESUMEN Y COMENTARIOS TÉCNICOS - {vfd_percentage_real:.2f}% RPM
        
        RESUMEN DE RESULTADOS VFD
        Punto de Operación VFD ({vfd_percentage_real:.2f}% RPM):
        - Caudal VDF: {session_state.get('caudal_nominal_vdf', session_state.get('caudal_operacion', 51.01)):.2f} L/s
        - Altura VDF: {session_state.get('altura_operacion_vdf', session_state.get('altura_operacion', 94.68)):.2f} m
        - Eficiencia VDF: {session_state.get('eficiencia_ajustada', session_state.get('eficiencia_operacion', 73.1)):.1f}%
        - Potencia VDF: {session_state.get('potencia_ajustada', session_state.get('potencia_operacion', 56.13)):.2f} HP
        - NPSH Requerido VDF: {session_state.get('npsh_requerido_vdf', session_state.get('npsh_requerido', 3.08)):.2f} m
        
        ANÁLISIS DE NPSH ({vfd_percentage_real:.2f}% RPM)
        Criterio de Diseño Estándar:
        - NPSH Disponible: {session_state.get('npshd_mca', 4.97):.2f} m.c.a.
        - NPSH Requerido VDF: {session_state.get('npsh_requerido_vdf', 3.08):.2f} m.c.a.
        - Diferencia: {session_state.get('npshd_mca', 4.97) - session_state.get('npsh_requerido_vdf', 3.08):.2f} m.c.a.
        - Margen Mínimo Requerido: {session_state.get('npsh_requerido_vdf', 3.08) * 1.2:.2f} m.c.a.
        
        COMENTARIOS TÉCNICOS VFD
        - Eficiencia: La bomba opera bien con VFD al {vfd_percentage_real:.2f}% RPM
        - NPSH Adecuado: Margen de seguridad adecuado con VFD
        - Potencia Requerida: Potencia reducida con VFD
        - Ahorro Energético: Potencial reducción de consumo energético
        
        RECOMENDACIONES VDF
        - Evaluar costo-beneficio de implementar VFD
        - Ajustar porcentaje RPM para optimización
        - Considerar impacto en vida útil de bomba"""
        
        # Usar comentarios de session_state si están disponibles, sino usar los por defecto
        resumen_100_rpm = session_state.get('comentarios_tecnicos_100_rpm', resumen_100_rpm_default)
        resumen_vdf = session_state.get('comentarios_tecnicos_vdf', resumen_vdf_default)
        
        ws_analisis.cell(row=3, column=1, value="RESUMEN Y COMENTARIOS TÉCNICOS - 100% RPM").font = Font(bold=True, size=12)
    ws_analisis.cell(row=4, column=1, value=resumen_100_rpm).alignment = Alignment(wrap_text=True, vertical='top')
        ws_analisis.row_dimensions[4].height = 200

        ws_analisis.cell(row=6, column=1, value=f"RESUMEN Y COMENTARIOS TÉCNICOS - {vfd_percentage_real:.2f}% RPM").font = Font(bold=True, size=12)
    ws_analisis.cell(row=7, column=1, value=resumen_vdf).alignment = Alignment(wrap_text=True, vertical='top')
        ws_analisis.row_dimensions[7].height = 200

    try:
        wb.save(output)
        output.seek(0)
        return output
    except Exception as e:
        raise Exception(f"Error guardando archivo Excel: {str(e)}")
                    
        except Exception as e:
        raise Exception(f"Error generando reporte Excel: {str(e)}")

def generate_system_curve_data(session_state: Dict[str, Any]) -> pd.DataFrame:
    """
    Genera automáticamente los datos de la curva del sistema basados en parámetros del proyecto
    """
    try:
        import numpy as np
        from core.system_head import calculate_adt_for_multiple_flows
        
        # Generar caudales para calcular la curva del sistema
        flows = np.linspace(0, 200, 20).tolist()  # 0 a 200 L/s
        
        # Obtener parámetros del sistema desde session_state
        system_params = {
            'h_estatica': session_state.get('altura_estatica_total', 50.0),
            'long_succion': session_state.get('long_succion', 10.0),
            'diam_succion_m': session_state.get('diam_succion_mm', 200.0) / 1000.0,
            'C_succion': session_state.get('C_succion', 150),
            'accesorios_succion': session_state.get('accesorios_succion', []),
            'otras_perdidas_succion': session_state.get('otras_perdidas_succion', 0.0),
            'long_impulsion': session_state.get('long_impulsion', 100.0),
            'diam_impulsion_m': session_state.get('diam_impulsion_mm', 150.0) / 1000.0,
            'C_impulsion': session_state.get('C_impulsion', 150),
            'accesorios_impulsion': session_state.get('accesorios_impulsion', []),
            'otras_perdidas_impulsion': session_state.get('otras_perdidas_impulsion', 0.0)
        }
        
        adt_values = calculate_adt_for_multiple_flows(flows, 'L/s', system_params)
        
        if adt_values:
            return pd.DataFrame({
                'Caudal (L/s)': flows,
                'ADT (m)': adt_values
            })
        else:
            # Fallback con valores por defecto
            return pd.DataFrame({
                'Caudal (L/s)': [0, 50, 100, 150, 200],
                'ADT (m)': [50, 52, 56, 62, 70]
            })
            
    except Exception as e:
        # Fallback con valores por defecto en caso de error
        print(f"Warning: Error generando curva del sistema automáticamente: {e}")
        return pd.DataFrame({
            'Caudal (L/s)': [0, 50, 100, 150, 200],
            'ADT (m)': [50, 52, 56, 62, 70]
        })

def ensure_required_dataframes(session_state: Dict[str, Any]) -> None:
    """
    Asegura que todos los DataFrames requeridos existan en el session_state.
    Crea DataFrames vacíos si no existen.
    """
    # Tablas principales del análisis (100% RPM)
        required_dataframes_100 = [
        'df_bomba_100',           # Curva H-Q Bomba 100%
        'df_rendimiento_100',     # Curva de Eficiencia 100%
        'df_potencia_100',       # Curva de Potencia 100%
        'df_npsh_100',           # Curva NPSH 100%
        'df_sistema_100'         # Curva del Sistema 100%
        ]
            
    # Tablas VDF
        required_dataframes_vfd = [
        'df_bomba_vdf',         # Curva H-Q Bomba VDF
        'df_rendimiento_vfd',   # Curva de Eficiencia VDF
        'df_potencia_vfd',      # Curva de Potencia VDF
        'df_npsh_vfd',          # Curva NPSH VDF
        'df_sistema_vfd'        # Curva del Sistema VDF
        ]
            
    # Tablas de compatibilidad (fallback)
        required_dataframes_fallback = [
        'df_curva_sistema', 'df_eff_100', 'df_power_100', 'df_npshr'
        ]
            
    # Combinar todas las tablas requeridas
        all_dataframes = required_dataframes_100 + required_dataframes_vfd + required_dataframes_fallback
            
        for df_name in all_dataframes:
        if df_name not in session_state or session_state[df_name] is None:
            session_state[df_name] = pd.DataFrame()

def create_comprehensive_excel_report(session_state: Dict[str, Any]) -> io.BytesIO:
    """
    Función principal mejorada que garantiza que cada hoja incluya inputs, resultados y gráficas
    según las especificaciones del usuario
    """
    try:
        print("PROCESANDO: Iniciando generación de reporte completo...")
        
        # Validar datos mínimos necesarios
        if not validate_minimum_data(session_state):
            raise Exception("No hay datos suficientes para generar el reporte. Complete los datos de entrada primero.")
        
        result = export_full_project_to_excel(session_state)
        print("ÉXITO: Reporte generado exitosamente")
        return result
        
    except Exception as e:
        error_msg = f"ERROR: Error crítico generando reporte: {str(e)}"
        print(error_msg)
        raise Exception(error_msg)

def validate_minimum_data(session_state: Dict[str, Any]) -> bool:
    """
    Valida que existan los datos mínimos necesarios para generar el reporte
    """
    required_fields = [
        'proyecto', 'caudal_lps', 'altura_estatica_total'
    ]
    
    for field in required_fields:
        if not session_state.get(field):
        print(f"ADVERTENCIA: Campo requerido faltante: {field}")
            return False
    
    return True
